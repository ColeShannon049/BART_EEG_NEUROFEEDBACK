"""  # Start/continue docstring
FINALBART_slowColor.py  # Execute statement

Balloon Analogue Risk Task (BART) with:  # Execute statement
- Integrated neurofeedback (NF) pipeline (theta → z-score → colour)  # Execute statement
- Practice-trial baseline  # Execute statement
- Optional simulation mode (no EEG required)  # Execute statement
- Optional debug HUD with live z-score graph  # Execute statement
- Balloon colour that only updates once per second to reduce ERP interference  # Execute statement
"""  # Start/continue docstring

from psychopy import visual, event, core, sound  # Import dependency
from psychopy.hardware import keyboard  # Import dependency
import random, csv, os, time, json, numpy as np  # Import dependency
import re  # Regex for BIDS/manifest parsing
# ---------------- BIDS-style output naming ----------------
# Matches LabRecorder-style conventions (e.g., sub-P001_ses-S032_task-Default_run-001_*.xdf)
# We write the BART outputs into:
#   bart_data/sub-<ID>/ses-<ID>/beh/sub-<ID>_ses-<ID>_task-<TASK>_run-<RUN>_beh.(csv|xlsx)
BIDS_ROOT = "bart_data"  # Set BIDS_ROOT
TASK_LABEL = "BART"     # change to "Default" if you want exact task label match to your XDFs

def _sanitize_label(s: str) -> str:  # Define function _sanitize_label
    """Keep only alphanumerics; remove spaces/symbols to keep filenames safe."""  # Start/continue docstring
    s = (s or "").strip()  # Set s
    out = []  # Set out
    for ch in s:  # Loop over items
        if ch.isalnum():  # Conditional branch
            out.append(ch)  # Execute statement
    return "".join(out)  # Return value from function

def _ensure_prefixed_padded(label: str, prefix: str, width: int = 3) -> str:  # Define function _ensure_prefixed_padded
    """  # Start/continue docstring
    Accepts inputs like '001' or 'P001' and returns 'P001' (padded to width for numeric part).  # Execute statement
    """  # Start/continue docstring
    label = _sanitize_label(label)  # Set label
    if not label:  # Conditional branch
        return f"{prefix}{'0'*width}"  # Return value from function

    up = label.upper()  # Set up
    if up.startswith(prefix.upper()):  # Conditional branch
        body = up[len(prefix):]  # Set body
    else:  # Fallback branch
        body = up  # Set body

    # If numeric, pad; otherwise leave as-is
    if body.isdigit():  # Conditional branch
        body = body.zfill(width)  # Set body
    return f"{prefix}{body}"  # Return value from function

def _text_entry(win, title: str, prompt: str, default: str = "") -> str:  # Define function _text_entry
    """  # Start/continue docstring
    Simple on-screen text entry (no GUI dependency).  # Execute statement
    - Enter/Return = confirm  # Execute statement
    - Backspace = delete  # Execute statement
    - Esc = quit experiment  # Execute statement
    """  # Start/continue docstring
    cur = default  # Set cur
    title_stim = visual.TextStim(win, text=title, pos=(0, 260), height=36, color=UI_TEXT_COLOR, bold=True)  # Set title_stim
    prompt_stim = visual.TextStim(win, text=prompt, pos=(0, 120), height=26, color=UI_TEXT_COLOR, wrapWidth=1000)  # Set prompt_stim
    box = visual.Rect(win, width=800, height=70, pos=(0, 20), lineColor="white", fillColor=None, lineWidth=2)  # Set box
    entry = visual.TextStim(win, text="", pos=(0, 20), height=34, color=UI_TEXT_COLOR)  # Set entry
    # Set hint
    hint = visual.TextStim(win, text="Type, then press ENTER.  (Esc to quit)", pos=(0, -260), height=20, color=UI_ACCENT_COLOR)

    # clear old keys
    event.clearEvents(eventType='keyboard')  # Execute statement

    while True:  # Loop while condition holds
        keys = event.getKeys()  # Set keys
        for k in keys:  # Loop over items
            if k in ("escape",):  # Conditional branch
                cleanup_and_exit(fh=None, send_final=False)  # Call cleanup_and_exit()
            if k in ("return", "num_enter", "enter"):  # Conditional branch
                return cur  # Return value from function
            if k in ("backspace", "delete"):  # Conditional branch
                cur = cur[:-1]  # Set cur
                continue  # Skip to next loop iteration
            if k in ("space", "tab"):  # Conditional branch
                # skip whitespace in IDs
                continue  # Skip to next loop iteration

            # allow digits + letters + dash/underscore if you want; keep alnum only in sanitize later
            if len(k) == 1 and (k.isalnum() or k in ("-", "_")):  # Conditional branch
                cur += k  # Execute statement

        title_stim.draw()  # Execute statement
        prompt_stim.draw()  # Execute statement
        box.draw()  # Execute statement
        entry.text = cur if cur else ""  # Execute statement
        entry.draw()  # Execute statement
        hint.draw()  # Execute statement
        safe_flip()  # Call safe_flip()
        core.wait(0.01)  # Execute statement

def prompt_bids_ids(win):  # Define function prompt_bids_ids
    """  # Start/continue docstring
    Prompts participant/session/run on the first screen.  # Execute statement
    Returns (sub_label, ses_label, run_label, bids_outdir).  # Call Returns()
    """  # Start/continue docstring
    # A short explainer page first
    intro_t = visual.TextStim(win, text="Participant Info", pos=(0, 240), height=40, color=UI_TEXT_COLOR, bold=True)  # Set intro_t
    intro_b = visual.TextStim(
        win,
        text=(
            "Before we begin, enter the participant and session IDs.\n"  # Execute statement
            "These will be used to name the output files in a BIDS-style format.\n\n"  # Execute statement
            "Example:\n"  # Execute statement
            "  sub-P001_ses-S032_task-BART_run-001_beh.csv\n"  # Execute statement
            "  sub-P001_ses-S032_task-BART_run-001_beh.xlsx\n\n"  # Execute statement
            "Press SPACE to continue."  # Execute statement
        ),
        pos=(0, 40),
        height=26,
        color=UI_TEXT_COLOR,
        wrapWidth=1000,
        alignText="left"  # Set alignText
    )
    intro_t.draw(); intro_b.draw(); safe_flip()  # Execute statement
    k = event.waitKeys(keyList=["space", "escape"])  # Set k
    if "escape" in k:  # Conditional branch
        cleanup_and_exit(fh=None, send_final=False)  # Call cleanup_and_exit()

    # Set sub_in
    sub_in = _text_entry(win, "Participant ID", "Enter participant number (e.g., 001 or P001):", default="001")
    ses_in = _text_entry(win, "Session ID", "Enter session number (e.g., 001 or S001):", default="001")  # Set ses_in
    run_in = _text_entry(win, "Run #", "Enter run number (usually 001):", default="001")  # Set run_in

    sub_label = _ensure_prefixed_padded(sub_in, "P", width=3)  # Set sub_label
    ses_label = _ensure_prefixed_padded(ses_in, "S", width=3)  # Set ses_label
    run_label = _sanitize_label(run_in)  # Set run_label
    if (not run_label) or (not run_label.isdigit()):  # Conditional branch
        run_label = "001"  # Set run_label
    run_label = run_label.zfill(3)  # Set run_label

    bids_outdir = os.path.join(BIDS_ROOT, f"sub-{sub_label}", f"ses-{ses_label}", "beh")  # Set bids_outdir
    os.makedirs(bids_outdir, exist_ok=True)  # Execute statement
    return sub_label, ses_label, run_label, bids_outdir  # Return value from function


# ----------------------------------------------------------------------
# TOP-LEVEL MODE FLAGS
# NOTE: Top-level switches that control whether NF comes from EEG, simulated input, or sham.
# ----------------------------------------------------------------------

SIMULATE_NF = False  # Set SIMULATE_NF
SHOW_NF_HUD = True  # Set SHOW_NF_HUD
SHAM_NF = True  # Set SHAM_NF


# ----------------------------------------------------------------------
# Manifest-based condition selection
# ----------------------------------------------------------------------
# If MANIFEST_FILENAME exists (same folder as this script or current working dir),
# the participant's ID (entered on the first screen) is matched to the manifest's
# `subject_id` column. The task will then automatically switch modes:
#   condition == "SHAM" -> SHAM_NF=True,  SIMULATE_NF=False
#   condition == "NF"   -> SHAM_NF=False, SIMULATE_NF=False (expects EEG NF_Z unless you choose SIMULATE_NF)
#
# This is useful for double-blind running: participants only enter their ID and
# the program silently selects the correct condition.
MANIFEST_FILENAME = "manifest.xlsx"
MANIFEST_SHEETNAME = None  # None = first sheet
MANIFEST_ID_COL = "subject_id"
MANIFEST_COND_COL = "condition"
MANIFEST_NAME_COL = "name"
MANIFEST_HILO_COL = "high/low"  # matches your manifest header exactly
MANIFEST_OVERRIDES_MODES = True

# Filled after ID entry
MANIFEST_ROW = {}
CONDITION_LABEL = ""  # e.g., 'NF' or 'SHAM'

def _norm_sub_id(x):
    """Normalize participant IDs so manifest lookups are robust."""
    if x is None:
        return ""
    s = str(x).strip()
    # allow 'sub-P001' or 'P001' or '001'
    s = re.sub(r"^sub-", "", s, flags=re.IGNORECASE)
    s = s.replace(" ", "")
    if s.isdigit():
        s = f"P{int(s):03d}"
    return s.upper()

def _find_manifest_path():
    """Find manifest in script folder first, then current working directory."""
    candidates = []
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        candidates.append(os.path.join(here, MANIFEST_FILENAME))
    except Exception:
        pass
    candidates.append(os.path.join(os.getcwd(), MANIFEST_FILENAME))
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def load_manifest_row_for_subject(subject_id):
    """Return dict(row) from manifest.xlsx for this subject_id (or {{}} if not found)."""
    path = _find_manifest_path()
    if not path:
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[MANIFEST_SHEETNAME] if (MANIFEST_SHEETNAME and MANIFEST_SHEETNAME in wb.sheetnames) else wb.worksheets[0]
        # headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [("" if h is None else str(h).strip()) for h in headers]
        col_map = {h: i+1 for i, h in enumerate(headers) if h}
        # locate required columns
        if MANIFEST_ID_COL not in col_map:
            # try case-insensitive fallback
            for h in list(col_map.keys()):
                if h.lower() == MANIFEST_ID_COL.lower():
                    col_map[MANIFEST_ID_COL] = col_map[h]
                    break
        if MANIFEST_ID_COL not in col_map:
            return {}
        sid_key = _norm_sub_id(subject_id)
        # scan rows
        for r in range(2, ws.max_row + 1):
            sid = ws.cell(row=r, column=col_map[MANIFEST_ID_COL]).value
            if _norm_sub_id(sid) == sid_key:
                row = {}
                for h, cidx in col_map.items():
                    row[h] = ws.cell(row=r, column=cidx).value
                return row
        return {}
    except Exception as e:
        print(f"⚠️ Manifest read failed: {e}")
        return {}

def apply_condition_from_manifest(subject_id):
    """Set SHAM_NF / SIMULATE_NF based on manifest row. Returns (row, label)."""
    global SHAM_NF, SIMULATE_NF, MANIFEST_ROW, CONDITION_LABEL
    MANIFEST_ROW = load_manifest_row_for_subject(subject_id)
    label = ""
    if MANIFEST_ROW:
        # pull condition
        cond = ""
        for k in (MANIFEST_COND_COL, MANIFEST_COND_COL.lower(), MANIFEST_COND_COL.upper()):
            if k in MANIFEST_ROW and MANIFEST_ROW[k] is not None:
                cond = str(MANIFEST_ROW[k]).strip()
                break
        cond_u = cond.upper()
        if cond_u in ("SHAM", "S", "0"):
            label = "SHAM"
            if MANIFEST_OVERRIDES_MODES:
                SHAM_NF = True
                SIMULATE_NF = False
        elif cond_u in ("NF", "NEUROFEEDBACK", "1"):
            label = "NF"
            # Always enforce true NF mode from manifest (prevents accidental SIM/SHAM carry-over)
            SHAM_NF = False
            SIMULATE_NF = False
        else:
            # unknown label; do not override
            label = cond.strip()
    CONDITION_LABEL = label
    return MANIFEST_ROW, CONDITION_LABEL


# --- UI mode toggle (comfort vs high-contrast) ---  # Choose UI style without changing task logic
NIGHT_MODE = True  # True = softer/lower-contrast UI to reduce eye strain; False = classic black/white high contrast


# ----------------------------------------------------------------------
# LSL IMPORTS
# ----------------------------------------------------------------------

from pylsl import StreamInfo, StreamOutlet  # Import dependency
try:  # Begin protected block (handle errors)
    from pylsl import StreamInlet, resolve_byprop as _resolve_byprop  # Import dependency

    def resolve_byprop(prop, val, timeout=1.5):  # Define function resolve_byprop
        return _resolve_byprop(prop, val, timeout=timeout)  # Return value from function

    LSL_OK = True  # Set LSL_OK

except Exception:  # Handle an error case
    LSL_OK = False  # Set LSL_OK
    StreamInlet = None  # Set StreamInlet

    def resolve_byprop(*a, **k):  # Define function resolve_byprop
        return []  # Return value from function

# ----------------------------------------------------------------------
# BART + NF PARAMETERS
# NOTE: Task parameters: trial counts, pump limits, colour thresholds, EEG/NF settings, and animation timing.
# ----------------------------------------------------------------------

PRACTICE_TRIALS = 5  # Set PRACTICE_TRIALS
N_TRIALS = 40  # Set N_TRIALS
REQUIRE_NF = False  # True = block until EEG stream is found (ignored in SIM/SHAM)
# If you're using internal SIMULATE_NF or SHAM_NF, we never require a live EEG stream.
if SIMULATE_NF or SHAM_NF:
    REQUIRE_NF = False


EEG_STREAM_NAME = "openvibeSignal"  # Set EEG_STREAM_NAME
FS = 512.0  # Set FS
THETA_BAND = (4.0, 8.0)  # Set THETA_BAND
WIN_S = 2.0  # Set WIN_S
WIN_SAMPLES = int(FS * WIN_S)  # Set WIN_SAMPLES

FRONTAL_IDXS = [5, 6]  # Set FRONTAL_IDXS

Z_HIGH = 0.3  # Set Z_HIGH
Z_LOW = -0.7  # Set Z_LOW

# --- Approx. isoluminant colours (DKL) ---
# We use DKL colour space with a constant luminance component (first value).
# True isoluminance requires monitor/participant calibration; this is a strong default.
BALLOON_COLOR_SPACE = 'rgb'  # iso-ish RGB palette (no DKL)
# DKL = [luminance, L-M, S-(L+M)]
DKL_LUM = 0.00
ISO_RED    = [ 0.80, -0.20, -0.20]
ISO_YELLOW = [ 0.60,  0.60, -0.48]
ISO_GREEN  = [-0.40,  0.50, -0.20]

# --- Approx "iso-ish" luminance matching (no per-monitor calibration required) ---
# PsychoPy's true isoluminance workflow requires monitor calibration.
# This helper rescales each RGB color to match the relative luminance of ISO_YELLOW.
def _srgb_to_linear(u):
    return u / 12.92 if u <= 0.04045 else ((u + 0.055) / 1.055) ** 2.4


def _linear_to_srgb(u):
    return 12.92 * u if u <= 0.0031308 else 1.055 * (u ** (1.0 / 2.4)) - 0.055


def _rel_luminance(rgb_m1_1):
    r, g, b = [(x + 1.0) * 0.5 for x in rgb_m1_1]
    r, g, b = _srgb_to_linear(r), _srgb_to_linear(g), _srgb_to_linear(b)
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def _match_luminance(rgb_m1_1, target_lum):
    r, g, b = [(x + 1.0) * 0.5 for x in rgb_m1_1]
    rr, gg, bb = _srgb_to_linear(r), _srgb_to_linear(g), _srgb_to_linear(b)
    lum = 0.2126 * rr + 0.7152 * gg + 0.0722 * bb
    if lum <= 1e-9:
        return [float(x) for x in rgb_m1_1]
    scale = target_lum / lum
    rr, gg, bb = rr * scale, gg * scale, bb * scale
    rr, gg, bb = min(max(rr, 0.0), 1.0), min(max(gg, 0.0), 1.0), min(max(bb, 0.0), 1.0)
    rr, gg, bb = _linear_to_srgb(rr), _linear_to_srgb(gg), _linear_to_srgb(bb)
    r2, g2, b2 = rr * 2.0 - 1.0, gg * 2.0 - 1.0, bb * 2.0 - 1.0
    return [float(r2), float(g2), float(b2)]


_ISO_TARGET_LUM = _rel_luminance(ISO_YELLOW)
ISO_RED = _match_luminance(ISO_RED, _ISO_TARGET_LUM)
ISO_YELLOW = _match_luminance(ISO_YELLOW, _ISO_TARGET_LUM)
ISO_GREEN = _match_luminance(ISO_GREEN, _ISO_TARGET_LUM)

# Fallback palette (also DKL; keeps luminance constant)
BALLOON_COLOURS = [
    ISO_RED,
    ISO_YELLOW,
    ISO_GREEN,
]

PUMPS_MAX = 30  # maximum pumps per balloon (updated)
POINTS_PER_PUMP = 1  # Set POINTS_PER_PUMP
ITI = 0.5  # Set ITI
FIXATION_BASELINE = 0.5  # Set FIXATION_BASELINE
PUMP_DELAY = 0.5  # Set PUMP_DELAY
SCREEN_SIZE = [1200, 800]  # Set SCREEN_SIZE

# ---------------- COMFORT / EYE-STRAIN SETTINGS ----------------
# NOTE: These reduce harsh luminance/contrast and sharp edges to make the task easier on the eyes.
# ----------------- COMFORT / UI STYLE (EYE STRAIN) -----------------  # UI tuning knobs for participant comfort
# NOTE: Set NIGHT_MODE=True to lower contrast and reduce harsh transients; set False for classic high-contrast lab UI.
if NIGHT_MODE:  # softer UI for comfort
    UI_BG_COLOR      = [-0.08, -0.08, -0.08]  # slightly-above-black background (reduces contrast fatigue)
    UI_TEXT_COLOR    = [ 0.85,  0.85,  0.85]  # softer white for on-screen text
    UI_ACCENT_COLOR  = 'lightskyblue'  # gentle accent color for prompts
    DOT_RADIUS_PX    = 8  # smaller dot (less visually sharp)
    DOT_COLOR        = [-0.35, -0.35, -0.35]  # slightly gray dot (less harsh than pure black)
    DOT_OPACITY      = 0.85  # slightly transparent dot (reduces edge harshness)
    FLASH_COLOR      = [ 0.55,  0.55,  0.55]  # gray flash instead of pure white
    FLASH_OPACITY    = 0.12  # lower flash intensity (fewer bright transients)
else:  # classic high-contrast UI
    UI_BG_COLOR      = [-1.0, -1.0, -1.0]  # true black background
    UI_TEXT_COLOR    = [ 1.0,  1.0,  1.0]  # true white text
    UI_ACCENT_COLOR  = 'lightskyblue'  # keep accent consistent
    DOT_RADIUS_PX    = 10  # slightly larger dot (classic)
    DOT_COLOR        = [-1.0, -1.0, -1.0]  # pure black dot
    DOT_OPACITY      = 1.0  # fully opaque dot
    FLASH_COLOR      = [ 1.0,  1.0,  1.0]  # white flash (classic)
    FLASH_OPACITY    = 0.18  # a bit stronger flash than NIGHT_MODE


BALLOON_START_RADIUS = 45.0  # Set BALLOON_START_RADIUS
BALLOON_GROWTH_FACTOR = 1.04  # Set BALLOON_GROWTH_FACTOR
BALLOON_GROWTH_ADD = 2.0  # Set BALLOON_GROWTH_ADD
BALLOON_MAX_RADIUS = 265.0  # Set BALLOON_MAX_RADIUS

PUMP_ANIM_SEC = 0.18  # Set PUMP_ANIM_SEC
POP_ANIM_SEC  = 0.14  # Set POP_ANIM_SEC
FADE_ANIM_SEC = 0.20  # Set FADE_ANIM_SEC

CHANCE_NO_POP = 0.03  # ~chance balloon never pops by PUMPS_MAX (max pumps is PUMPS_MAX)

ALLOW_NEGATIVE_BANK = True  # Set ALLOW_NEGATIVE_BANK

Z_ALPHA = 0.6  # Set Z_ALPHA

NF_UPDATE_HZ = 10.0  # Set NF_UPDATE_HZ
NF_UPDATE_INTERVAL = 1.0 / NF_UPDATE_HZ  # Set NF_UPDATE_INTERVAL
NF_COLOR_UPDATE_INTERVAL = 1.0  # Set NF_COLOR_UPDATE_INTERVAL


COLOR_FADE_SEC = 0.30  # Fade balloon color changes over 200–400ms to reduce luminance transients / ERP interference
NF_STUCK_TIMEOUT = 1.0  # Set NF_STUCK_TIMEOUT
NF_FLAT_STEPS = 50  # Set NF_FLAT_STEPS
NF_FLAT_EPS = 1e-3  # Set NF_FLAT_EPS

BOOM_DUR    = 1.0  # Set BOOM_DUR
COLLECT_DUR = 0.8  # Set COLLECT_DUR
EASE_IN  = 0.12  # Set EASE_IN
EASE_OUT = 0.30  # Set EASE_OUT

FLASH_DUR = 0.05  # shorter flash to reduce eye strain and luminance transients

# ----------------- REST END CHIME -----------------
# Short tone played at the END of eyes-closed rest blocks (EC) to cue participants to open eyes.
CHIME_ENABLED = True  # Set CHIME_ENABLED
END_CHIME_HZ  = 880.0  # Set END_CHIME_HZ
END_CHIME_DUR = 0.30   # seconds
END_CHIME_VOL = 0.25   # 0..1 (keep modest to reduce startling)

try:  # Begin protected block (handle errors)
    end_chime = sound.Sound(value=END_CHIME_HZ, secs=END_CHIME_DUR, stereo=True)  # Set end_chime
    end_chime.setVolume(END_CHIME_VOL)  # Execute statement
    CHIME_OK = True  # Set CHIME_OK
except Exception as _e:  # Handle an error case
    print("⚠️ Chime sound init failed (audio unavailable):", _e)  # Print debug/status message
    end_chime = None  # Set end_chime
    CHIME_OK = False  # Set CHIME_OK

GREEN_SUCCESS_FRAC   = 0.60  # Set GREEN_SUCCESS_FRAC
GREEN_STREAK_TARGET  = 3  # Set GREEN_STREAK_TARGET
BONUS_POINTS         = 10  # Set BONUS_POINTS

DEBUG_GRAPH = True  # Set DEBUG_GRAPH
GRAPH_WIDTH  = 320  # Set GRAPH_WIDTH
GRAPH_HEIGHT = 160  # Set GRAPH_HEIGHT
GRAPH_POS    = (-420, -300)  # Set GRAPH_POS
GRAPH_Z_RANGE = 2.5  # Set GRAPH_Z_RANGE

# ----------------------------------------------------------------------
# WINDOW
# ----------------------------------------------------------------------

win = visual.Window(
    size=SCREEN_SIZE,
    units='pix',
    color=UI_BG_COLOR,  # softer background to reduce eye strain
    fullscr=False,
    allowGUI=False,
    # NOTE (lab GPU stability): some Windows lab machines render Circle stimuli as a
    # full-screen solid color when Frame Buffer Objects (FBOs) are enabled.
    # Disabling FBOs fixes the "balloon color fills the whole screen" symptom.
    useFBO=True,
    winType='pyglet',
    waitBlanking=True,
    checkTiming=False,
)

def safe_flip():  # Define function safe_flip
    try:  # Begin protected block (handle errors)
        win.flip()  # Execute statement
        return True  # Return value from function
    except Exception as e:  # Handle an error case
        print("Flip error:", e)  # Print debug/status message
        return False  # Return value from function

# ----------------------------------------------------------------------
# LSL MARKER STREAM
# ----------------------------------------------------------------------

info = StreamInfo(
    name="BART_Markers",
    type="Markers",
    channel_count=1,
    nominal_srate=0,
    channel_format="string",
    source_id="bart_psychopy",
)
outlet = StreamOutlet(info)  # Set outlet

def send_marker(code: str, **data):  # Define function send_marker
    t = core.getTime()  # Set t
    meta = ";".join([f"{k}={v}" for k, v in data.items()])  # Set meta
    msg = f"{code};timestamp={t};{meta}"  # Set msg
    try:  # Begin protected block (handle errors)
        outlet.push_sample([msg])  # Execute statement
    except Exception:  # Handle an error case
        pass  # No-op placeholder

def cleanup_and_exit(fh=None, send_final=True, total_bank=0):  # Define function cleanup_and_exit
    """Close files/window safely. Also writes an XLSX copy of the CSV data."""  # Start/continue docstring
    try:  # Begin protected block (handle errors)
        if send_final:  # Conditional branch
            try:  # Begin protected block (handle errors)
                outlet.push_sample([f"BART_END;timestamp={core.getTime()};total={total_bank}"])  # Execute statement
            except Exception:  # Handle an error case
                pass  # No-op placeholder

        # Try XLSX export before closing (uses rows_buffer accumulated during the run)
        try:  # Begin protected block (handle errors)
            if 'xlsxfile' in globals() and 'rows_buffer' in globals() and 'FIELDNAMES' in globals():  # Conditional branch
                write_xlsx(xlsxfile, rows_buffer, FIELDNAMES)  # Call write_xlsx()
        except Exception as e:  # Handle an error case
            print("⚠️ XLSX export failed:", e)  # Print debug/status message

        if fh:  # Conditional branch
            try:  # Begin protected block (handle errors)
                fh.flush()  # Execute statement
                fh.close()  # Execute statement
            except Exception:  # Handle an error case
                pass  # No-op placeholder

        try:  # Begin protected block (handle errors)
            win.close()  # Execute statement
        except Exception:  # Handle an error case
            pass  # No-op placeholder
    finally:  # Run cleanup regardless of errors
        raise SystemExit(0)  # Raise exception to stop/handle flow

# ----------------------------------------------------------------------
# VISUAL STIMULI
# ----------------------------------------------------------------------

trial_text = visual.TextStim(win, pos=(0, 360), height=22, color=UI_TEXT_COLOR)  # Set trial_text
total_text = visual.TextStim(win, pos=(420, 360), height=20, color=UI_TEXT_COLOR)  # Set total_text

balloon = visual.Circle(
    win,
    radius=BALLOON_START_RADIUS,
    edges=128,
    fillColor=ISO_YELLOW,
    lineColor=None,
)

pump_dot = visual.Circle(
    win,
    radius=DOT_RADIUS_PX,  # smaller dot to reduce eye strain
    edges=48,  # still smooth, slightly less visually 'sharp'
    lineColor=DOT_COLOR,  # slightly gray (less harsh than black)
    fillColor=DOT_COLOR,  # slightly gray (less harsh than black)
    opacity=DOT_OPACITY,  # soften the dot edge further
    pos=(0, 0)  # centered inside the balloon
)

pump_value_text = visual.TextStim(
    win,
    text="",
    pos=(0, 0),
    height=20,
    color=UI_TEXT_COLOR,
    bold=True  # Set bold
)

fixation = visual.TextStim(win, text="+", height=64, color=UI_TEXT_COLOR)  # Set fixation

boom_text = visual.TextStim(win, text="BOOM!", height=44, color=UI_TEXT_COLOR, pos=(0, 24), bold=True)  # Set boom_text
loss_text = visual.TextStim(win, text="", height=28, color=UI_TEXT_COLOR, pos=(0, -14), bold=True)  # Set loss_text
collect_text = visual.TextStim(win, text="", height=44, color=UI_TEXT_COLOR, pos=(0, 0), bold=True)  # Set collect_text

note_text = visual.TextStim(win, text="", pos=(0, -120), height=24, color="black")  # Set note_text

nf_status = visual.TextStim(
    win, pos=(-560, 360), height=16,
    color=UI_ACCENT_COLOR  # Set color
)

bonus_text_main = visual.TextStim(win, text="", height=40, color=UI_TEXT_COLOR, pos=(0, 20))  # Set bonus_text_main
bonus_text_sub  = visual.TextStim(win, text="", height=24, color=UI_TEXT_COLOR, pos=(0, -30))  # Set bonus_text_sub

flash_rect = visual.Rect(
    win,
    width=SCREEN_SIZE[0],
    height=SCREEN_SIZE[1],
    fillColor=FLASH_COLOR,  # gray flash to reduce luminance transient
    opacity=0.0,
)

# ----------------------------------------------------------------------
# BALLOON RESET (robust across GPUs)
# ----------------------------------------------------------------------

def reset_balloon_visual(fill_color=None):
    """Reset balloon back to default state (without triggering full-screen fill)."""

    # Stop any leftover animations from the previous trial
    try:
        inflate_tween.active = False
        pop_tween.active = False
        fade_tween.active = False
    except Exception:
        pass

    # Reset color if requested
    if fill_color is not None:
        # copy so we don't accidentally share references
        balloon.fillColor = list(fill_color)

    # Reset shape state
    balloon.opacity = 1.0
    balloon.radius = float(BALLOON_START_RADIUS)

    # IMPORTANT: do NOT set balloon.size here (can double-scale on some machines)




graph_frame = visual.Rect(
    win,
    width=GRAPH_WIDTH,
    height=GRAPH_HEIGHT,
    lineColor=[0.3, 0.3, 0.3],
    pos=GRAPH_POS,
)

graph_line = visual.ShapeStim(
    win,
    vertices=[(0, 0), (1, 0)],
    lineColor=UI_ACCENT_COLOR,
    closeShape=False,
)

graph_zero = visual.Line(
    win,
    start=(GRAPH_POS[0] - GRAPH_WIDTH / 2, GRAPH_POS[1]),
    end=(GRAPH_POS[0] + GRAPH_WIDTH / 2, GRAPH_POS[1]),
    lineColor=[0.2, 0.2, 0.2],
)

# ----------------------------------------------------------------------
# TRIAL SETUP
# ----------------------------------------------------------------------

def make_trials(n):  # Define function make_trials
    cols = (BALLOON_COLOURS * (n // len(BALLOON_COLOURS) + 1))[:n]  # Set cols
    random.shuffle(cols)  # Execute statement
    return [{"trial_num": i + 1, "colour": cols[i]} for i in range(n)]  # Return value from function

practice_trials = make_trials(PRACTICE_TRIALS)
# Force practice balloons to start as the neutral/yellow balloon for comfort + consistency
for _tr in practice_trials:
    _tr['colour'] = ISO_YELLOW
  # Set practice_trials
main_trials     = make_trials(N_TRIALS)  # Set main_trials

# ---------------- OUTPUT FILES (BIDS-style) ----------------
# Prompt participant/session/run BEFORE anything else that writes files
SUB_LABEL, SES_LABEL, RUN_LABEL, outdir = prompt_bids_ids(win)  # Execute statement

# Determine condition from manifest (silent switch to SHAM/NF)
MANIFEST_ROW, CONDITION_LABEL = apply_condition_from_manifest(SUB_LABEL)
if CONDITION_LABEL:
    print(f"🧾 Manifest condition for sub-{SUB_LABEL}: {CONDITION_LABEL}")
else:
    print(f"🧾 No manifest match/label for sub-{SUB_LABEL} (continuing with current mode flags).")

bids_base = f"sub-{SUB_LABEL}_ses-{SES_LABEL}_task-{TASK_LABEL}_run-{RUN_LABEL}_beh"  # Set bids_base
csvfile  = os.path.join(outdir, bids_base + ".csv")  # Set csvfile
xlsxfile = os.path.join(outdir, bids_base + ".xlsx")  # Set xlsxfile

# (Optional) keep a timestamped backup copy inside the same folder
timestamp = time.strftime("%Y%m%d-%H%M%S")  # Set timestamp

f = open(csvfile, "w", newline="")  # Set f
writer = csv.DictWriter(
    f,
    fieldnames=[
                "sub",
                "ses",
                "run",
                "task",
                "subject_id",
                "name",
                "high/low",
                "condition",
                "block",
        "trial",
        "colour",
        "pump_count",
        "exploded",
        "explosion_point",
        "trial_value",
        "loss_if_pop",
        "trial_earnings",
        "total_earnings",
        "events",
        "trial_start_time",
        "trial_end_time",
        "trial_duration_sec",
        "nf_source",
        "z_used",
        "nf_color",
        "nf_green_frac",
        "nf_green_success",
        "adjusted_pumps_trial",
        "exploded_int",
        "collected",
        "pump_latency_first",
        "pump_latency_mean",
        "pump_latency_median",
        "collect_latency_from_ready",
        "collect_latency_from_trial_start",
        "pump_latencies_json",
        "pump_times_json",
        "baseline_mu",
        "baseline_sigma",
        "baseline_n",
        "rest_pre_eo_theta_mean",
        "rest_pre_eo_theta_std",
        "rest_pre_eo_z_mean",
        "rest_pre_eo_z_std",
        "rest_pre_eo_n",
        "rest_pre_ec_theta_mean",
        "rest_pre_ec_theta_std",
        "rest_pre_ec_z_mean",
        "rest_pre_ec_z_std",
        "rest_pre_ec_n",
        "rest_pre_conc_theta_mean",
        "rest_pre_conc_theta_std",
        "rest_pre_conc_z_mean",
        "rest_pre_conc_z_std",
        "rest_pre_conc_n",
        "rest_post_eo_theta_mean",
        "rest_post_eo_theta_std",
        "rest_post_eo_z_mean",
        "rest_post_eo_z_std",
        "rest_post_eo_n",
        "rest_post_ec_theta_mean",
        "rest_post_ec_theta_std",
        "rest_post_ec_z_mean",
        "rest_post_ec_z_std",
        "rest_post_ec_n"  # Execute statement
    ],
)
writer.writeheader()  # Execute statement

# Keep rows in memory for optional XLSX export
FIELDNAMES = list(writer.fieldnames)  # Set FIELDNAMES
rows_buffer = []  # each element is a dict row written to CSV

def _safe_str(v):  # Define function _safe_str
    """Convert values to something Excel/openpyxl can write.  # Start/continue docstring

    openpyxl cannot write numpy scalars/arrays or Python containers directly.  # Execute statement
    We convert:  # Execute statement
      - numpy scalars -> float/int  # Execute statement
      - list/tuple/np.ndarray -> compact "r,g,b" if length==3 numeric else JSON string  # Execute statement
      - dict -> JSON string  # Execute statement
      - everything else -> as-is  # Execute statement
    """  # Start/continue docstring
    if v is None:  # Conditional branch
        return ""  # Return value from function
    # numpy scalar
    try:  # Begin protected block (handle errors)
        import numpy as _np  # Import dependency
        if isinstance(v, _np.generic):  # Conditional branch
            return v.item()  # Return value from function
        if isinstance(v, _np.ndarray):  # Conditional branch
            v = v.tolist()  # Set v
    except Exception:  # Handle an error case
        pass  # No-op placeholder

    # containers
    if isinstance(v, dict):  # Conditional branch
        try:  # Begin protected block (handle errors)
            import json as _json  # Import dependency
            return _json.dumps(v)  # Return value from function
        except Exception:  # Handle an error case
            return str(v)  # Return value from function

    if isinstance(v, (list, tuple)):  # Conditional branch
        # special-case RGB triplets for readability
        if len(v) == 3 and all(isinstance(x, (int, float)) for x in v):  # Conditional branch
            return ",".join([f"{float(x):.5g}" for x in v])  # Return value from function
        try:  # Begin protected block (handle errors)
            import json as _json  # Import dependency
            return _json.dumps(v)  # Return value from function
        except Exception:  # Handle an error case
            return str(v)  # Return value from function

    # fall back
    return v  # Return value from function

def write_xlsx(xlsx_path, rows, fieldnames):  # Define function write_xlsx
    """Write trial-level rows + a small summary sheet to an .xlsx file."""  # Start/continue docstring
    try:  # Begin protected block (handle errors)
        from openpyxl import Workbook  # Import dependency
    except Exception as e:  # Handle an error case
        print("⚠️ openpyxl not available; cannot write XLSX:", e)  # Print debug/status message
        return False  # Return value from function

    wb = Workbook()  # Set wb
    ws = wb.active  # Set ws
    ws.title = "trials"  # Execute statement

    # Header
    ws.append(fieldnames)  # Execute statement

    # Rows
    for r in rows:  # Loop over items
        ws.append([_safe_str(r.get(k, "")) for k in fieldnames])  # Execute statement

    # Summary sheet (per block)
    try:  # Begin protected block (handle errors)
        ws2 = wb.create_sheet("summary")  # Set ws2
        ws2.append(["block",
                    "n_trials",
                    "explosion_frequency",
                    "mean_adjusted_pumps",
                    "mean_pump_latency",
                    "median_pump_latency",
                    "final_total_earnings"])  # Execute statement
        blocks = {}  # Set blocks
        for r in rows:  # Loop over items
            b = r.get("block", "")  # Set b
            blocks.setdefault(b, []).append(r)  # Execute statement

        for b, rs in blocks.items():  # Loop over items
            n = len(rs)  # Set n
            exploded_vals = [int(r.get("exploded_int", 0) or 0) for r in rs]  # Set exploded_vals
            explosion_freq = (sum(exploded_vals) / n) if n else ""  # Set explosion_freq

            adj = []  # Set adj
            for r in rs:  # Loop over items
                ap = r.get("adjusted_pumps_trial", "")  # Set ap
                if ap == "" or ap is None:  # Conditional branch
                    continue  # Skip to next loop iteration
                try:  # Begin protected block (handle errors)
                    adj.append(float(ap))  # Execute statement
                except Exception:  # Handle an error case
                    pass  # No-op placeholder
            mean_adj = (sum(adj) / len(adj)) if adj else ""  # Set mean_adj

            lat_mean = []  # Set lat_mean
            lat_med = []  # Set lat_med
            for r in rs:  # Loop over items
                v1 = r.get("pump_latency_mean", "")  # Set v1
                v2 = r.get("pump_latency_median", "")  # Set v2
                try:  # Begin protected block (handle errors)
                    if v1 != "" and v1 is not None:  # Conditional branch
                        lat_mean.append(float(v1))  # Execute statement
                except Exception:  # Handle an error case
                    pass  # No-op placeholder
                try:  # Begin protected block (handle errors)
                    if v2 != "" and v2 is not None:  # Conditional branch
                        lat_med.append(float(v2))  # Execute statement
                except Exception:  # Handle an error case
                    pass  # No-op placeholder

            mean_lat = (sum(lat_mean) / len(lat_mean)) if lat_mean else ""  # Set mean_lat
            # median across trial medians
            if lat_med:  # Conditional branch
                lat_med_sorted = sorted(lat_med)  # Set lat_med_sorted
                mid = len(lat_med_sorted)//2  # Set mid
                median_lat = (lat_med_sorted[mid] if len(lat_med_sorted)%2==1  # Set median_lat
                              else (lat_med_sorted[mid-1]+lat_med_sorted[mid])/2.0)  # Call else()
            else:  # Fallback branch
                median_lat = ""  # Set median_lat

            # final total earnings: last row's total for that block
            final_total = ""  # Set final_total
            try:  # Begin protected block (handle errors)
                final_total = rs[-1].get("total_earnings", "")  # Set final_total
            except Exception:  # Handle an error case
                pass  # No-op placeholder

            ws2.append([b, n, explosion_freq, mean_adj, mean_lat, median_lat, final_total])  # Execute statement
    except Exception as e:  # Handle an error case
        print("⚠️ Could not write summary sheet:", e)  # Print debug/status message


    # ---------------- Pump-level sheet (one row per pump) ----------------
    try:  # Begin protected block (handle errors)
        import json as _json  # Import dependency
        ws_p = wb.create_sheet("pumps")  # Set ws_p
        pump_header = [
            "block", "trial",
            "pump_number",
            "pump_latency_sec",
            "pump_time_sec",
            "exploded_int",
            "collected",
            "explosion_point",
            "trial_value",
            "trial_earnings",
            "total_earnings",
            "nf_source",
            "z_used",
            "nf_color",
        ]
        ws_p.append(pump_header)  # Execute statement

        for r in rows:  # Loop over items
            block = r.get("block", "")  # Set block
            trial = r.get("trial", "")  # Set trial
            exploded_int = int(r.get("exploded_int", 0) or 0)  # Set exploded_int
            collected = int(r.get("collected", 0) or 0)  # Set collected
            exp_point = r.get("explosion_point", "")  # Set exp_point
            trial_value = r.get("trial_value", "")  # Set trial_value
            trial_earn = r.get("trial_earnings", "")  # Set trial_earn
            total_earn = r.get("total_earnings", "")  # Set total_earn
            nf_source = r.get("nf_source", "")  # Set nf_source
            z_used = r.get("z_used", "")  # Set z_used
            nf_color = r.get("nf_color", "")  # Set nf_color

            lat_j = r.get("pump_latencies_json", "") or ""  # Set lat_j
            t_j   = r.get("pump_times_json", "") or ""  # Set t_j

            try:  # Begin protected block (handle errors)
                lat_list = _json.loads(lat_j) if lat_j else []  # Set lat_list
            except Exception:  # Handle an error case
                lat_list = []  # Set lat_list
            try:  # Begin protected block (handle errors)
                t_list = _json.loads(t_j) if t_j else []  # Set t_list
            except Exception:  # Handle an error case
                t_list = []  # Set t_list

            n_p = max(len(lat_list), len(t_list))  # Set n_p
            for i in range(n_p):  # Loop over items
                lat = lat_list[i] if i < len(lat_list) else ""  # Set lat
                tt  = t_list[i] if i < len(t_list) else ""  # Set tt
                ws_p.append([
                    block, trial,
                    i + 1,
                    lat,
                    tt,
                    exploded_int,
                    collected,
                    exp_point,
                    trial_value,
                    trial_earn,
                    total_earn,
                    nf_source,
                    z_used,
                    nf_color,
                ])
    except Exception as e:  # Handle an error case
        print("⚠️ Could not write pump-level sheet:", e)  # Print debug/status message

    try:  # Begin protected block (handle errors)
        wb.save(xlsx_path)  # Execute statement
        print("✅ XLSX saved:", xlsx_path)  # Print debug/status message
        return True  # Return value from function
    except Exception as e:  # Handle an error case
        print("⚠️ Could not save XLSX:", e)  # Print debug/status message
        return False  # Return value from function


# ----------------------------------------------------------------------
# NF COLOR LOGIC
# ----------------------------------------------------------------------

def z_to_color(z):  # Define function z_to_color
    """  # Start/continue docstring
    Map z-score → (balloon_colour, category_string).  # Execute statement
    - z >= Z_HIGH → GREEN  # Execute statement
    - z <= Z_LOW  → RED  # Execute statement
    - otherwise   → YELLOW  # Execute statement
    """  # Start/continue docstring
    if z is None:  # Conditional branch
        return ISO_YELLOW, "mid"  # Return value from function
    if z >= Z_HIGH:  # Conditional branch
        return ISO_GREEN, "high"  # Return value from function
    if z <= Z_LOW:  # Conditional branch
        return ISO_RED, "low"  # Return value from function
    return ISO_YELLOW, "mid"  # Return value from function


# Sham deterministic signal (~4800 samples)
SHAM_THETA = (
    [
        0.42, 0.45, 0.48, 0.52, 0.55, 0.58, 0.62, 0.66, 0.61, 0.57,
        0.53, 0.49, 0.46, 0.44, 0.43, 0.45, 0.48, 0.52, 0.56, 0.61,
        0.65, 0.64, 0.60, 0.55, 0.50, 0.47, 0.46, 0.48, 0.50, 0.53,
        0.57, 0.62, 0.66, 0.63, 0.59, 0.54, 0.50, 0.47, 0.46, 0.47,
    ]
    * 120  # Execute statement
)


# ----------------------------------------------------------------------
# NF CONNECTOR
# NOTE: Neurofeedback engine: connects to EEG (or sim/sham), computes theta power, builds baseline, outputs z-score, and keeps history for the HUD graph.
# ----------------------------------------------------------------------


class NFConnector:
    """Neurofeedback connector.

    Supports three modes:
      1) Real NF: read z-scores from LSL stream named 'NF_Z' (type 'NF').
      2) SIMULATE_NF: generate an internal simulated theta + z (no EEG/LSL needed).
      3) SHAM_NF: deterministic fake theta + z (no EEG/LSL needed).

    Rest baselines (EO/EC/Concentrated) are used to estimate baseline μ/σ and
    a *direction* sign so that 'better' always maps to positive z.

    Notes:
    - In real NF mode, we only receive z (unless you also stream theta). We still
      allow the task to run; rest-theta summaries may be blank if no theta is available.
    """

    def __init__(self):
        self.inlet = None
        self.connected = False

        # Latest values
        self.last_z = 0.0
        self.ema = 0.0  # EMA-smoothed z
        self.last_theta = None
        self.last_theta_time = None

        # Baseline params
        self.baseline_done = False
        self.baseline_vals = []  # baseline sample count for HUD
        self.baseline_mu = 0.0
        self.baseline_sigma = 1.0
        self.baseline_direction = 1.0
        self.baseline_method = ''
        self.baseline_n = 0
        self.baseline_rest_anchor = float('nan')
        self.baseline_conc_anchor = float('nan')

        # SIM state
        self._sim_theta = 0.0
        self._sim_t = 0.0

        # SHAM state
        self._sham_idx = 0
        # --- debug histories for HUD/graph (safe on all machines) ---
        self.history_z = []      # rolling z-score history
        self.history_theta = []  # rolling theta (or proxy) history
        self._hist_maxlen = 300  # ~30s at 10Hz (adjust as needed)
        self.warning_text = ''  # optional HUD warning line

    def try_connect(self, attempts=10, sleep_s=0.5):
        """Try to connect to LSL 'NF_Z'. If SIM/SHAM is enabled, no connection is needed."""
        if SIMULATE_NF or SHAM_NF:
            self.connected = True
            return True
        if not LSL_OK:
            return False
        for _ in range(attempts):
            streams = resolve_byprop('name', 'NF_Z', timeout=1.0)
            if not streams:
                streams = resolve_byprop('type', 'NF', timeout=1.0)
            if streams:
                try:
                    self.inlet = StreamInlet(streams[0], max_buflen=120, recover=True)
                    self.connected = True
                    return True
                except Exception:
                    self.inlet = None
                    self.connected = False
            core.wait(sleep_s)
        return False

    def _sim_step(self):
        """One step of simulated theta (random-walk with gentle mean reversion)."""
        # Mean-reverting random walk around 0
        noise = random.uniform(-1.0, 1.0)
        self._sim_theta += 0.10*noise - 0.01*self._sim_theta
        self._sim_t = core.getTime()
        self.last_theta = float(self._sim_theta)
        self.last_theta_time = self._sim_t


    def _sham_step(self):
        """Realistic SHAM neurofeedback generator (z-like).

        Produces piecewise-stable LOW/MID/HIGH streaks:
          • Mostly stable segments (reduces rapid colour switching)
          • Occasional long HIGH segments (~20s) to mimic 'good' runs
          • Small within-segment variability so it doesn't feel deterministic

        The output is used exactly like NF z-scores: it drives balloon colour via
        Z_LOW / Z_HIGH thresholds.
        """
        import random, math, time
        now = time.perf_counter()

        # init
        if (not getattr(self, "_sham_seeded", False)) or (getattr(self, "_sham_last_t", None) is None):
            self._sham_state = 'mid'
            self._sham_z = -0.15
            self._sham_until = now + random.uniform(4.0, 8.0)
            self._sham_seeded = True
            self._sham_last_t = now

        dt = max(1e-3, now - (self._sham_last_t or now))
        self._sham_last_t = now

        # state transition when streak ends
        if now >= self._sham_until:
            cur = self._sham_state

            # transition probabilities (stay common, big jumps rare)
            if cur == 'low':
                nxt = random.choices(['low','mid','high'], weights=[0.50, 0.30, 0.20], k=1)[0]
            elif cur == 'mid':
                nxt = random.choices(['low','mid','high'], weights=[0.30, 0.40, 0.30], k=1)[0]
            else:  # high
                nxt = random.choices(['low','mid','high'], weights=[0.25, 0.30, 0.45], k=1)[0]

            self._sham_state = nxt

            # durations: mix short + occasional long highs
            if random.random() < 0.33:
                dur = random.uniform(18.0, 28.0)   # long streak
            elif random.random() < 0.65:
                dur = random.uniform(12.0, 16.0)   # medium streak
            else:
                dur = random.uniform(5.0, 8.0)     # short streak

            self._sham_until = now + dur

        targets = {'low': -1.05, 'mid': -0.10, 'high': 0.60}
        target = targets.get(self._sham_state, -0.10)

        # smooth mean-reverting dynamics (OU-like)
        tau = 1.8
        sigma = 0.06
        self._sham_z += (target - self._sham_z) * (dt / tau) + sigma * math.sqrt(dt) * random.gauss(0.0, 1.0)

        # clamp around target to avoid flicker across thresholds
        band = 0.22
        lo = target - band
        hi = target + band
        if self._sham_z < lo: self._sham_z = lo
        if self._sham_z > hi: self._sham_z = hi

        return float(self._sham_z)


    def set_baseline_from_rest_epochs(self, rest_eo_samples, conc_samples):
        """Compute baseline from dedicated rest blocks.

        Robust approach:
          - μ = median(rest EO theta)
          - σ = 1.4826*MAD(rest EO theta) (fallback to std)
          - direction = +1 if conc median >= rest median else -1

        If there are too few samples, we fall back to a neutral baseline.
        """
        import numpy as np

        rest = np.asarray(rest_eo_samples, float)
        rest = rest[np.isfinite(rest)]
        conc = np.asarray(conc_samples, float)
        conc = conc[np.isfinite(conc)]

        if len(rest) < 5:
            self.baseline_mu = 0.0
            self.baseline_sigma = 1.0
            self.baseline_direction = 1.0
            self.baseline_done = True
            self.baseline_method = 'rest_fallback_neutral'
            self.baseline_n = int(len(rest))
            print('[NF] Rest baseline insufficient → using neutral baseline (μ=0, σ=1).')
            return

        mu = float(np.median(rest))
        mad = float(np.median(np.abs(rest - mu)))
        sigma = 1.4826*mad if mad > 0 else float(np.std(rest, ddof=1) if len(rest) > 1 else 1.0)
        if sigma == 0.0:
            sigma = 1.0

        if len(conc) >= 5:
            conc_mu = float(np.median(conc))
            direction = 1.0 if conc_mu >= mu else -1.0
        else:
            conc_mu = float('nan')
            direction = 1.0

        # Optional anchors (for logs)
        def _mean_top(arr, frac=0.10):
            if len(arr) == 0:
                return float('nan')
            k = max(1, int(round(frac*len(arr))))
            return float(np.mean(np.sort(arr)[-k:]))

        def _mean_bottom(arr, frac=0.10):
            if len(arr) == 0:
                return float('nan')
            k = max(1, int(round(frac*len(arr))))
            return float(np.mean(np.sort(arr)[:k]))

        self.baseline_mu = mu
        self.baseline_sigma = sigma
        self.baseline_direction = direction
        self.baseline_done = True
        self.baseline_method = 'rest_EO_median_MAD + conc_direction'
        self.baseline_n = int(len(rest))
        self.baseline_rest_anchor = _mean_top(rest, 0.10)
        self.baseline_conc_anchor = _mean_bottom(conc, 0.10) if len(conc) else float('nan')

        print(f"[NF] Baseline set: μ={mu:.3e} σ={sigma:.3e} dir={direction:+.0f} n={len(rest)} conc_med={conc_mu:.3e}")

    def pull_z(self):
        """Return the current z-score (EEG/LSL, SIM, or SHAM) and update debug histories."""
        # timing
        now_t = core.getTime()
        if getattr(self, "_last_pull_t", None) is None:
            dt = 0.1
        else:
            dt = max(1e-3, float(now_t - self._last_pull_t))
        self._last_pull_t = now_t

        # Make sure histories exist for HUD/graph
        if not hasattr(self, "history_z"):
            self.history_z = []
        if not hasattr(self, "history_theta"):
            self.history_theta = []
        if not hasattr(self, "_hist_maxlen"):
            self._hist_maxlen = 300

        def _push_hist(z, theta=None):
            try:
                self.history_z.append(float(z))
                if theta is not None:
                    self.history_theta.append(float(theta))
                if len(self.history_z) > self._hist_maxlen:
                    self.history_z = self.history_z[-self._hist_maxlen:]
                if len(self.history_theta) > self._hist_maxlen:
                    self.history_theta = self.history_theta[-self._hist_maxlen:]
            except Exception:
                pass

        # ---------- SHAM ----------
        if SHAM_NF:
            # Piecewise "streak" sham (NOT sinusoidal): holds LOW/MID/HIGH for sampled durations.
            # Designed to feel plausible and include a few longer GREEN streaks (~20s) mixed with quicker changes.
            z = self._sham_step(dt)
            # Update EMA and histories for HUD/graph
            self.last_z = float(z)
            self.ema = Z_ALPHA * self.last_z + (1.0 - Z_ALPHA) * self.ema
            _push_hist(self.ema, self.last_theta)
            return self.ema

        # ---------- EEG/LSL ----------

        if (not self.connected) or (self.inlet is None):
            _push_hist(self.ema)
            return self.ema

        chunk = None
        try:
            chunk, _ = self.inlet.pull_chunk(timeout=0.0, max_samples=32)
        except Exception:
            chunk = None

        if chunk:
            try:
                self.last_z = float(chunk[-1][0])
            except Exception:
                pass

        self.ema = Z_ALPHA * float(self.last_z) + (1.0 - Z_ALPHA) * self.ema
        _push_hist(self.ema)
        return self.ema

    def _push_hist(self, z, theta=None):
        """Store rolling history for debug HUD/graph; safe if lists are missing."""
        try:
            if not hasattr(self, 'history_z'):
                self.history_z = []
            if not hasattr(self, 'history_theta'):
                self.history_theta = []
            if not hasattr(self, '_hist_maxlen'):
                self._hist_maxlen = 300
            self.history_z.append(float(z))
            if theta is not None:
                self.history_theta.append(float(theta))
            if len(self.history_z) > self._hist_maxlen:
                self.history_z = self.history_z[-self._hist_maxlen:]
            if len(self.history_theta) > self._hist_maxlen:
                self.history_theta = self.history_theta[-self._hist_maxlen:]
        except Exception:
            pass

    def _mean_top(arr, frac=0.10):
        arr = np.asarray(arr, float)
        arr = arr[np.isfinite(arr)]
        if len(arr) == 0:
            return float('nan')
        k = max(1, int(round(frac * len(arr))))
        return float(np.mean(np.sort(arr)[-k:]))

    def _mean_bottom(arr, frac=0.10):
        arr = np.asarray(arr, float)
        arr = arr[np.isfinite(arr)]
        if len(arr) == 0:
            return float('nan')
        k = max(1, int(round(frac * len(arr))))
        return float(np.mean(np.sort(arr)[:k]))

    # ------------------------------------------------------------------
    # Theta calculation (simulation / sham / real)
    # ------------------------------------------------------------------
    def _compute_theta_power(self):  # Define function _compute_theta_power
        # SHAM deterministic sequence
        if SHAM_NF:  # Conditional branch
            val = SHAM_THETA[self.sham_index]  # Set val
            self.sham_index = (self.sham_index + 1) % len(SHAM_THETA)  # Execute statement
            self.last_theta = float(val)  # Execute statement
            self.last_theta_time = core.getTime()  # Execute statement
            return self.last_theta  # Return value from function

        # Simulated NF
        if SIMULATE_NF:  # Conditional branch
            self.sim_t += 1.0 / NF_UPDATE_HZ  # Execute statement
            base = 1.0 + 0.2 * np.sin(0.3 * self.sim_t)  # Set base
            noise = 0.05 * np.random.randn()  # Set noise

            # Arrow key drift
            keys = event.getKeys(keyList=["up", "down"])  # Set keys
            if "up" in keys:  # Conditional branch
                self.sim_drift += 0.02  # Execute statement
            if "down" in keys:  # Conditional branch
                self.sim_drift -= 0.02  # Execute statement
            self.sim_drift = float(np.clip(self.sim_drift, -1.0, 1.0))  # Execute statement

            theta = base + self.sim_drift + noise  # Set theta
            self.last_theta = float(theta)  # Execute statement
            self.last_theta_time = core.getTime()  # Execute statement
            return self.last_theta  # Return value from function

        # Real EEG LSL mode
        if not (self.connected and self.eeg_inlet and self.buffer is not None):  # Conditional branch
            return None  # Return value from function

        chunk, _ = self.eeg_inlet.pull_chunk(timeout=0.0, max_samples=WIN_SAMPLES)  # Execute statement
        if not chunk:  # Conditional branch
            return None  # Return value from function

        arr = np.asarray(chunk)  # Set arr
        if arr.ndim != 2 or arr.shape[1] <= max(FRONTAL_IDXS):  # Conditional branch
            return None  # Return value from function

        n_new = min(arr.shape[0], WIN_SAMPLES)  # Set n_new
        new_data = arr[-n_new:, FRONTAL_IDXS]  # Set new_data

        self.buffer = np.roll(self.buffer, -n_new, axis=0)  # Execute statement
        self.buffer[-n_new:, :] = new_data  # Execute statement

        data = self.buffer.T.astype(float)  # Set data
        data -= data.mean(axis=1, keepdims=True)  # Execute statement

        fft_vals = np.fft.rfft(data, axis=1)  # Set fft_vals
        freqs = np.fft.rfftfreq(data.shape[1], 1.0 / FS)  # Set freqs

        mask = (freqs >= THETA_BAND[0]) & (freqs <= THETA_BAND[1])  # Set mask
        if not np.any(mask):  # Conditional branch
            return None  # Return value from function

        psd = np.abs(fft_vals) ** 2  # Set psd
        theta_power = float(psd[:, mask].mean())  # Set theta_power

        self.last_theta = theta_power  # Execute statement
        self.last_theta_time = core.getTime()  # Execute statement
        return theta_power  # Return value from function

    # ------------------------------------------------------------------
    # History
    # ------------------------------------------------------------------
    def _append_history(self, z):  # Define function _append_history
        self.history_z.append(float(z))  # Execute statement
        if len(self.history_z) > self.history_len:  # Conditional branch
            self.history_z = self.history_z[-self.history_len:]  # Execute statement

    # ------------------------------------------------------------------
    # Public update
    # ------------------------------------------------------------------
    def pull_z(self):  # Define function pull_z
        now = core.getTime()  # Set now

        # try connection if not connected
        if not self.connected:  # Conditional branch
            self.try_connect(attempts=1)  # Execute statement
            if not self.connected:  # Conditional branch
                self.last_z = 0.0  # Execute statement
                self._append_history(self.last_z)  # Execute statement
                return self.last_z  # Return value from function

        # decimation
        theta = None  # Set theta
        if (now - self.last_update_time) >= NF_UPDATE_INTERVAL:  # Conditional branch
            self.last_update_time = now  # Execute statement
            theta = self._compute_theta_power()  # Set theta

        if theta is None:  # Conditional branch
            self._append_history(self.last_z)  # Execute statement
            return self.last_z  # Return value from function

        if self.baseline_active and not self.baseline_done:  # Conditional branch
            self.baseline_vals.append(theta)  # Execute statement

        # compute z
        if self.baseline_done:  # Conditional branch
            z_raw = (theta - self.baseline_mu) / self.baseline_sigma  # Set z_raw
        else:  # Fallback branch
            z_raw = 0.0  # Set z_raw

        # EMA smoothing
        if self.ema is None:  # Conditional branch
            self.ema = z_raw  # Execute statement
        else:  # Fallback branch
            self.ema = Z_ALPHA * z_raw + (1.0 - Z_ALPHA) * self.ema  # Execute statement

        self.last_z = self.ema  # Execute statement

        self._append_history(self.last_z)  # Execute statement
        return self.last_z  # Return value from function


# ----------------------------------------------------------------------
# NFConnector pull_z SAFEGUARD (ROBUST)
#
# Why this exists:
#   - On some lab machines (or after merges), parts of the EEG/theta-compute code
#     may not be available (e.g., missing _compute_theta_power).
#   - We still want *SIM* and *SHAM* to work with no EEG, and we want the HUD graph
#     to always have data.
#
# What this does:
#   - SHAM_NF: emits a deterministic z pattern (updates every NF_UPDATE_INTERVAL)
#   - SIMULATE_NF: emits an internal z (random walk) (also supports UP/DOWN nudges)
#   - EEG/LSL: reads z directly from the NF_Z LSL stream when available
#
# NOTE: This *overrides* NFConnector.pull_z unconditionally so the task never breaks.
# ----------------------------------------------------------------------

def _nf_pull_z_safeguard(self):
    """Return current z-score in all modes and keep history for the HUD graph."""
    now = core.getTime()

    # Ensure history exists (for HUD graph)
    if not hasattr(self, 'history_z'):
        self.history_z = []
    if not hasattr(self, 'history_len'):
        self.history_len = 240

    def _append(zval: float):
        self.history_z.append(float(zval))
        if len(self.history_z) > int(self.history_len):
            self.history_z[:] = self.history_z[-int(self.history_len):]

    # ---------------- SHAM MODE ----------------
    if SHAM_NF:
        if not hasattr(self, 'last_update_time'):
            self.last_update_time = 0.0
        if (now - float(self.last_update_time)) >= float(NF_UPDATE_INTERVAL):
            self.last_update_time = now
            pattern = [-0.8, -0.4, 0.0, 0.4, 0.8, 0.4, 0.0, -0.4]
            idx = int(getattr(self, 'sham_index', 0)) % len(pattern)
            self.sham_index = idx + 1
            self.last_z = float(pattern[idx])
        z = float(getattr(self, 'last_z', 0.0))
        _append(z)
        return z

    # ---------------- SIM MODE ----------------
    if SIMULATE_NF:
        if not hasattr(self, 'last_update_time'):
            self.last_update_time = 0.0
        if not hasattr(self, 'sim_z'):
            self.sim_z = 0.0

        # Optional: allow gentle manual nudging with UP/DOWN (doesn't affect BART keys)
        ks = event.getKeys(keyList=['up', 'down'])
        if 'up' in ks:
            self.sim_z += 0.15
        if 'down' in ks:
            self.sim_z -= 0.15

        if (now - float(self.last_update_time)) >= float(NF_UPDATE_INTERVAL):
            self.last_update_time = now
            # random-walk drift
            self.sim_z += random.gauss(0.0, 0.08)
            self.sim_z = max(-3.0, min(3.0, self.sim_z))
            self.last_z = float(self.sim_z)

        z = float(getattr(self, 'last_z', 0.0))
        _append(z)
        return z

    # ---------------- EEG/LSL MODE ----------------
    # Read z directly from NF_Z stream (do NOT compute theta here)
    if not getattr(self, 'connected', False):
        try:
            self.try_connect(attempts=1)
        except Exception:
            pass

    z_raw = None
    if getattr(self, 'inlet', None) is not None:
        try:
            chunk, _ts = self.inlet.pull_chunk(timeout=0.0, max_samples=16)
            if chunk:
                z_raw = float(chunk[-1][0])
        except Exception:
            z_raw = None

    if z_raw is None:
        z = float(getattr(self, 'last_z', 0.0))
        _append(z)
        return z

    # EMA smoothing
    ema = getattr(self, 'ema', None)
    if ema is None:
        ema = z_raw
    else:
        ema = float(Z_ALPHA) * float(z_raw) + (1.0 - float(Z_ALPHA)) * float(ema)
    self.ema = ema
    self.last_z = float(ema)

    z = float(self.last_z)
    _append(z)
    return z


# Override pull_z unconditionally so SIM/SHAM never depend on EEG/theta code.
if 'NFConnector' in globals():
    NFConnector.pull_z = _nf_pull_z_safeguard


# ---------------- NFConnector: REST-BASELINE METHOD PATCH ----------------
# The previous auto-generated versions sometimes left a nested function definition
# instead of a real NFConnector method. This patch ensures the method exists.

def _nf_set_baseline_from_rest_epochs(self, rest_eo_samples, conc_samples, rest_ec_samples=None):
    """Compute NF baseline from dedicated rest epochs.

    Inputs (lists of floats):
      - rest_eo_samples: theta power samples from Eyes-Open Rest
      - conc_samples: theta power samples from Concentrated Rest
      - rest_ec_samples: optional Eyes-Closed Rest samples (logged, not required)

    Robust baseline:
      μ = median(rest EO)
      σ = 1.4826 * MAD(rest EO)  (falls back to std if MAD too small)
      direction = +1 if conc median >= rest median else -1
        (keeps "higher z" aligned with the intended 'more theta in concentration' assumption)

    Sets:
      self.baseline_mu, self.baseline_sigma, self.baseline_direction,
      self.baseline_done, self.baseline_method, self.baseline_n,
      self.baseline_rest_anchor, self.baseline_conc_anchor, self.baseline_ec_anchor
    """
    import numpy as np

    rest = np.asarray(rest_eo_samples, float)
    rest = rest[np.isfinite(rest)]
    conc = np.asarray(conc_samples, float)
    conc = conc[np.isfinite(conc)]
    ec = np.asarray(rest_ec_samples, float) if rest_ec_samples is not None else np.asarray([], float)
    ec = ec[np.isfinite(ec)]

    # defaults if insufficient data
    mu = 0.0
    sigma = 1.0
    direction = 1.0

    if len(rest) >= 10:
        mu = float(np.median(rest))
        mad = float(np.median(np.abs(rest - mu)))
        sigma = float(1.4826 * mad)
        if not np.isfinite(sigma) or sigma < 1e-12:
            sigma = float(np.std(rest, ddof=1)) if len(rest) >= 2 else 1.0
        if not np.isfinite(sigma) or sigma < 1e-12:
            sigma = 1.0

        conc_mu = float(np.median(conc)) if len(conc) else mu
        direction = 1.0 if conc_mu >= mu else -1.0
    else:
        conc_mu = float(np.median(conc)) if len(conc) else 0.0

    def _mean_top(arr, frac):
        if len(arr) == 0:
            return float('nan')
        k = max(1, int(round(frac * len(arr))))
        return float(np.mean(np.sort(arr)[-k:]))

    def _mean_bottom(arr, frac):
        if len(arr) == 0:
            return float('nan')
        k = max(1, int(round(frac * len(arr))))
        return float(np.mean(np.sort(arr)[:k]))

    self.baseline_mu = mu
    self.baseline_sigma = sigma
    self.baseline_direction = direction
    self.baseline_done = True
    self.baseline_method = "rest_EO_median_MAD + conc_direction"
    self.baseline_n = int(len(rest))

    # helpful anchors for logging/debugging
    self.baseline_rest_anchor = _mean_top(rest, 0.10) if len(rest) else float('nan')
    self.baseline_conc_anchor = _mean_bottom(conc, 0.10) if len(conc) else float('nan')
    self.baseline_ec_anchor = float(np.median(ec)) if len(ec) else float('nan')

    try:
        print(f"[NF] Baseline set: μ={mu:.3e} σ={sigma:.3e} dir={direction:+.0f} n={len(rest)} conc_med={conc_mu:.3e}")
    except Exception:
        pass

# Attach as a real method on the class
try:
    NFConnector.set_baseline_from_rest_epochs = _nf_set_baseline_from_rest_epochs
except Exception:
    pass


# ----------------------------------------------------------------------
# TWEEN CLASS FOR SMOOTH ANIMATION
# ----------------------------------------------------------------------

class Tween:  # Define class Tween
    """  # Start/continue docstring
    Simple scalar tween:  # Execute statement
    - start at v0, end at v1  # Execute statement
    - over duration 'dur' (seconds)  # Execute statement
    - cubic ease-out for a smooth visual feel  # Execute statement
    """  # Start/continue docstring
    def __init__(self):  # Define function __init__
        self.active = False  # Execute statement
        self.t0 = 0.0  # Execute statement
        self.dur = 0.0  # Execute statement
        self.v0 = 0.0  # Execute statement
        self.v1 = 0.0  # Execute statement
        self.value = 0.0  # Execute statement

    def start(self, v0, v1, dur):  # Define function start
        self.active = True  # Execute statement
        self.t0 = time.perf_counter()  # Execute statement
        self.dur = max(1e-6, float(dur))  # Execute statement
        self.v0 = float(v0)  # Execute statement
        self.v1 = float(v1)  # Execute statement
        self.value = self.v0  # Execute statement

    def update(self):  # Define function update
        if not self.active:  # Conditional branch
            return self.value  # Return value from function
        t = (time.perf_counter() - self.t0) / self.dur  # Set t
        if t >= 1.0:  # Conditional branch
            self.value = self.v1  # Execute statement
            self.active = False  # Execute statement
            return self.value  # Return value from function
        eased = 1.0 - (1.0 - t) ** 3  # cubic ease-out
        self.value = self.v0 + (self.v1 - self.v0) * eased  # Execute statement
        return self.value  # Return value from function


inflate_tween = Tween()  # Set inflate_tween
pop_tween = Tween()  # Set pop_tween
fade_tween = Tween()  # Set fade_tween

# ----------------------------------------------------------------------
# EXPLOSION HAZARD (LINEAR HAZARD + "NO POP" OUTCOME)
# ----------------------------------------------------------------------

def draw_explosion_point_linear(pmax: int, p_no_pop: float):  # Define function draw_explosion_point_linear
    """  # Start/continue docstring
    Draw an explosion pump index using a linear hazard (weights ∝ 1..pmax).  # Execute statement
    If a random draw falls under p_no_pop, returns None (no explosion).  # Execute statement
    """  # Start/continue docstring
    if random.random() < p_no_pop:  # Conditional branch
        return None  # Return value from function
    ks = list(range(1, pmax + 1))  # Set ks
    return random.choices(ks, weights=ks, k=1)[0]  # Return value from function

# ----------------------------------------------------------------------
# KEYBOARD + HUD
# ----------------------------------------------------------------------

kb = keyboard.Keyboard()  # Set kb

def draw_hud(block, num, total, pumps, bank):  # Define function draw_hud
    """  # Start/continue docstring
    Draw minimal HUD: trial counter + bank.  # Execute statement
    (Pump count / value can be added back if desired.)  # Execute statement
    """  # Start/continue docstring
    trial_text.text = f"{block} {num}/{total}"  # Execute statement
    total_text.text = f"Bank: {bank} pts"  # Execute statement

    trial_text.draw()  # Execute statement
    total_text.draw()  # Execute statement

# ----------------------------------------------------------------------
# OVERLAY EASING HELPERS
# ----------------------------------------------------------------------

def ease_io(t, dur, ein=EASE_IN, eout=EASE_OUT):  # Define function ease_io
    """  # Start/continue docstring
    Ease-in / hold / ease-out envelope for opacity (0..1) over duration dur.  # Execute statement
    """  # Start/continue docstring
    if dur <= 0:  # Conditional branch
        return 1.0  # Return value from function
    if t <= 0:  # Conditional branch
        return 0.0  # Return value from function
    if t >= dur:  # Conditional branch
        return 0.0  # Return value from function

    if t < ein:  # Conditional branch
        x = t / ein  # Set x
        return x * x * (3 - 2 * x)  # Return value from function

    if t < (dur - eout):  # Conditional branch
        return 1.0  # Return value from function

    x = (t - (dur - eout)) / eout  # Set x
    x = min(max(x, 0.0), 1.0)  # Set x
    return 1.0 - (x * x * (3 - 2 * x))  # Return value from function

def float_up(y0, y1, t, dur):  # Define function float_up
    """  # Start/continue docstring
    Smooth vertical float from y0 to y1 over dur (cubic ease-out).  # Execute statement
    """  # Start/continue docstring
    if dur <= 0:  # Conditional branch
        return y1  # Return value from function
    tt = min(max(t / dur, 0.0), 1.0)  # Set tt
    return y0 + (y1 - y0) * (1.0 - (1.0 - tt) ** 3)  # Return value from function

# ----------------------------------------------------------------------
# DEBUG GRAPH DRAWING
# ----------------------------------------------------------------------

def draw_debug_graph(nf: NFConnector):  # Define function draw_debug_graph
    if not DEBUG_GRAPH or not SHOW_NF_HUD:  # Conditional branch
        return  # Return value from function
    if not getattr(nf, 'history_z', []):  # Conditional branch
        return  # Return value from function

    graph_frame.draw()  # Execute statement
    graph_zero.draw()  # Execute statement

    z_arr = np.asarray(getattr(nf, 'history_z', []), dtype=float)  # Set z_arr
    n = len(z_arr)  # Set n
    xs = np.linspace(-GRAPH_WIDTH / 2, GRAPH_WIDTH / 2, n)  # Set xs

    z_clip = np.clip(z_arr, -GRAPH_Z_RANGE, GRAPH_Z_RANGE)  # Set z_clip
    ys = (z_clip / GRAPH_Z_RANGE) * (GRAPH_HEIGHT / 2)  # Set ys

    verts = list(zip(xs + GRAPH_POS[0], ys + GRAPH_POS[1]))  # Set verts
    graph_line.vertices = verts  # Execute statement
    graph_line.draw()  # Execute statement

# ----------------------------------------------------------------------
# SINGLE TRIAL LOGIC (BART)
# NOTE: Core BART trial loop: reads keys, pumps/collects, triggers explosion, logs latencies, updates balloon colour from NF once per second, and writes per-trial rows.
# ----------------------------------------------------------------------

def run_trial(block_name, idx, trial, n_in_block, bank, nf: NFConnector, use_fallback_colors=False, nf_color_enabled=True):  # Define function run_trial
    """  # Start/continue docstring
    Run one BART balloon trial.  # Execute statement

    Returns:  # Execute statement
        (new_bank, nf_green_success)  # Execute statement
    """  # Start/continue docstring
    trial_start = core.getTime()  # Set trial_start
    tnum = trial["trial_num"]  # Set tnum

    last_color_update = core.getTime()  # Set last_color_update

    nf_cat = ""  # Set nf_cat
    if nf.connected and nf_color_enabled:  # Conditional branch
        z = nf.pull_z()  # Set z
        col, cat = z_to_color(z)  # Execute statement
        balloon.fillColor = col  # Execute statement
        nf_cat = cat  # Set nf_cat

        theta_txt = f"{nf.last_theta:.3e}" if nf.last_theta is not None else "n/a"  # Set theta_txt
        mu_txt = f"{nf.baseline_mu:.2e}" if nf.baseline_mu is not None else "n/a"  # Set mu_txt
        sig_txt = f"{nf.baseline_sigma:.2e}" if nf.baseline_sigma is not None else "n/a"  # Set sig_txt
        n_b = len(getattr(nf,'baseline_vals',[]))  # Set n_b
        main_line = (
            f"NF {'SHAM' if SHAM_NF else ('SIM' if SIMULATE_NF else 'EEG')}: θ={theta_txt}  z={z:.2f}  "  # Execute statement
            f"μ={mu_txt} σ={sig_txt} n={n_b}  "  # Execute statement
            f"[baseline {'OK' if nf.baseline_done else '...'}]"  # Execute statement
        )
        if nf.warning_text:  # Conditional branch
            nf_status.text = main_line + "\n" + nf.warning_text  # Execute statement
            nf_status.color = "orange"  # Execute statement
        else:  # Fallback branch
            nf_status.text = main_line  # Execute statement
            nf_status.color = "lightskyblue"  # Execute statement

        nf_src = "SHAM_NF" if SHAM_NF else ("SIM" if SIMULATE_NF else "EEG")  # Set nf_src
        z_used = z  # Set z_used
    else:  # Fallback branch
        col = trial["colour"] if use_fallback_colors else ISO_YELLOW  # Set col
        balloon.fillColor = col  # Execute statement
        nf_status.text = "NF: NONE (no EEG / sim)"  # Execute statement
        nf_status.color = "lightskyblue"  # Execute statement
        nf_src = "NONE"  # Set nf_src
        z_used = ""  # Set z_used

    reset_balloon_visual()  # Reset size/opacity/color for new trial
    balloon_visible = True  # Set balloon_visible

    # ----------------- COLOR FADE (ERP-FRIENDLY) -----------------
    # We avoid sudden luminance/color steps by fading balloon color updates over ~300 ms.
    # We also *freeze* any color transitions on the exact explosion frame (marker frame).
    color_fade_active = False  # True while we interpolate fillColor → targetColor
    color_fade_t0 = 0.0  # Start time (core.getTime) for the current color fade
    color_fade_from = list(balloon.fillColor)  # Starting RGB triplet for fade
    color_fade_to = list(balloon.fillColor)  # Target RGB triplet for fade
    freeze_color_until = 0.0  # When > now, we do NOT change balloon color (e.g., explosion marker frame)

    def _ease_in_out(u):  # Smoothstep easing for color fades
        u = 0.0 if u < 0.0 else (1.0 if u > 1.0 else u)  # Clamp to [0,1]
        return u * u * (3.0 - 2.0 * u)  # Smoothstep

    def start_color_fade(new_col, dur=COLOR_FADE_SEC):  # Begin a new color fade toward new_col
        nonlocal color_fade_active, color_fade_t0, color_fade_from, color_fade_to  # Use outer-scope vars
        color_fade_from = list(balloon.fillColor)  # Fade from the CURRENT displayed color
        color_fade_to = list(new_col)  # Fade toward the NEW target color
        color_fade_t0 = core.getTime()  # Record start time
        color_fade_active = True  # Arm fade

    def update_color_fade(now):  # Apply the active fade (if any) to balloon.fillColor
        nonlocal color_fade_active  # Update active flag in outer scope
        if not color_fade_active:  # If no fade is active, do nothing
            return  # Exit
        dur = max(1e-6, float(COLOR_FADE_SEC))  # Avoid divide-by-zero
        u = (now - color_fade_t0) / dur  # Fraction elapsed
        if u >= 1.0:  # Fade complete
            balloon.fillColor = color_fade_to  # Snap exactly to target
            color_fade_active = False  # Disarm fade
            return  # Exit
        u = _ease_in_out(u)  # Ease fraction for smoother transitions
        balloon.fillColor = [  # Interpolate each RGB channel
            color_fade_from[0] + (color_fade_to[0] - color_fade_from[0]) * u,
            color_fade_from[1] + (color_fade_to[1] - color_fade_from[1]) * u,
            color_fade_from[2] + (color_fade_to[2] - color_fade_from[2]) * u,
        ]

    pumps = 0  # Set pumps
    exploded = False  # Set exploded
    earnings = 0  # Set earnings

    explosion_point = draw_explosion_point_linear(PUMPS_MAX, CHANCE_NO_POP)  # Set explosion_point
    events = [f"hazard=linear;pNoPop={CHANCE_NO_POP}"]  # Set events

    cool_until = 0.0  # Set cool_until
    collect_until = 0.0  # Set collect_until
    boom_until = 0.0  # Set boom_until
    flash_until = 0.0  # Set flash_until
    max_reached = False  # Set max_reached

    collected = False  # Set collected
    show_dot = True  # Set show_dot

    # --- latency logging (behavioral outputs) ---
    pump_latencies = []   # seconds: 'ready' cue (dot visible) -> SPACE press
    pump_times = []       # seconds since trial start, for each pump press

    # --- overlay playback helpers (GPU/driver-robust) ---
    # On some lab PCs, alpha blending / draw order can make brief overlays appear to "never show".
    # These helpers render BOOM/collect overlays in a dedicated mini-loop so they always get time on-screen.
    def play_collect_animation(earnings_pts: int):
        """Show +points overlay for a fixed duration, with eased opacity/position."""
        t0 = core.getTime()
        while True:
            now_anim = core.getTime()
            t = now_anim - t0
            if t >= COLLECT_DUR:
                break

            # Draw stable scene (no dot during overlay)
            if balloon_visible:
                balloon.draw()
            draw_hud(block_name, tnum, n_in_block, pumps, bank)

            # NF HUD / debug (optional)
            if SHOW_NF_HUD:
                nf_status.draw()
                if DEBUG_GRAPH:
                    draw_debug_graph(nf)

            # Overlay text
            op = ease_io(t, COLLECT_DUR)
            yc = float_up(-6, 10, t, COLLECT_DUR)
            collect_text.opacity = op
            collect_text.pos = (0, yc)
            collect_text.draw()

            safe_flip()
            core.wait(0.005)

        # Reset opacity so next trial starts clean on all GPUs
        collect_text.opacity = 1.0

    def play_boom_animation(loss_pts: int):
        """Show BOOM + loss overlay for a fixed duration, with flash behind."""
        t0 = core.getTime()
        while True:
            now_anim = core.getTime()
            t = now_anim - t0
            if t >= BOOM_DUR:
                break

            # Flash behind everything (short pulse at the start)
            if t < FLASH_DUR:
                flash_rect.opacity = FLASH_OPACITY
                flash_rect.draw()
            else:
                flash_rect.opacity = 0.0

            # No balloon/dot on explosion overlay (reduces ERP-unrelated transients)
            draw_hud(block_name, tnum, n_in_block, pumps, bank)

            if SHOW_NF_HUD:
                nf_status.draw()
                if DEBUG_GRAPH:
                    draw_debug_graph(nf)

            op = ease_io(t, BOOM_DUR)
            y_main = float_up(16, 28, t, BOOM_DUR)
            y_sub  = float_up(-24, -10, t, BOOM_DUR)
            boom_text.text = "BOOM!"
            loss_text.text = f"-{loss_pts} pts"
            boom_text.opacity = op
            loss_text.opacity = op
            boom_text.pos = (0, y_main)
            loss_text.pos = (0, y_sub)
            boom_text.draw()
            loss_text.draw()

            safe_flip()
            core.wait(0.005)

        # Reset
        flash_rect.opacity = 0.0
        boom_text.opacity = 1.0
        loss_text.opacity = 1.0
    last_ready_time = core.getTime()  # dot is visible on first frame
    collect_latency_from_ready = ""  # Set collect_latency_from_ready
    collect_latency_from_trial_start = ""  # Set collect_latency_from_trial_start
    nf_frames = 0  # Set nf_frames
    nf_high_frames = 0  # Set nf_high_frames

    boom_t0 = None  # Set boom_t0
    collect_t0 = None  # Set collect_t0

    balloon.draw()  # Execute statement
    draw_hud(block_name, tnum, n_in_block, pumps, bank)  # Call draw_hud()
    if show_dot:  # Conditional branch
        pump_dot.draw()  # Execute statement
    if SHOW_NF_HUD:  # Conditional branch
        nf_status.draw()  # Execute statement
    if DEBUG_GRAPH and SHOW_NF_HUD:  # Conditional branch
        draw_debug_graph(nf)  # Call draw_debug_graph()
    safe_flip()  # Call safe_flip()

    send_marker(
        "BART_TRIAL_START",
        block=block_name,
        trial=tnum,
        expoint=(explosion_point if explosion_point is not None else -1),
        nf=nf_src,
        z=(round(z_used, 3) if isinstance(z_used, (float, int)) else ""),
    )

    # ---------------------- MAIN TRIAL LOOP ----------------------
    while True:  # Loop while condition holds
        now = core.getTime()  # Set now

        # ----------------- NF update (theta + z) -----------------
        if nf.connected and nf_color_enabled:  # Conditional branch
            z = nf.pull_z()  # Set z

            theta_txt = f"{nf.last_theta:.3e}" if nf.last_theta is not None else "n/a"  # Set theta_txt
            mu_txt = f"{nf.baseline_mu:.2e}" if nf.baseline_mu is not None else "n/a"  # Set mu_txt
            sig_txt = f"{nf.baseline_sigma:.2e}" if nf.baseline_sigma is not None else "n/a"  # Set sig_txt
            n_b = len(getattr(nf,'baseline_vals',[]))  # Set n_b
            main_line = (
                f"NF {'SHAM' if SHAM_NF else ('SIM' if SIMULATE_NF else 'EEG')}: θ={theta_txt}  z={z:.2f}  "  # Execute statement
                f"μ={mu_txt} σ={sig_txt} n={n_b}  "  # Execute statement
                f"[baseline {'OK' if nf.baseline_done else '...'}]"  # Execute statement
            )
            if nf.warning_text:  # Conditional branch
                nf_status.text = main_line + "\n" + nf.warning_text  # Execute statement
                nf_status.color = "orange"  # Execute statement
            else:  # Fallback branch
                nf_status.text = main_line  # Execute statement
                nf_status.color = "lightskyblue"  # Execute statement

            if (now - last_color_update) >= NF_COLOR_UPDATE_INTERVAL:  # Conditional branch
                # Update the *target* color only once per second, then fade smoothly to it.
                # This reduces sudden luminance transients that can contaminate ERP timing.
                if now >= freeze_color_until and not exploded and now >= boom_until:  # Guard: never change color on explosion marker frame
                    col, cat = z_to_color(z)  # Compute target color from z-score
                    start_color_fade(col)  # Fade toward new color (instead of instant change)
                    nf_cat = cat  # Track categorical NF state (high/mid/low)
                last_color_update = now  # Update timer regardless (keeps cadence stable)

            nf_frames += 1  # Execute statement
            if nf_cat == "high":  # Conditional branch
                nf_high_frames += 1  # Execute statement

        # Apply any smooth balloon-color fade (ERP-friendly).  #
        # Note: we freeze color changes during explosion marker frames.  #
        if now >= freeze_color_until and not exploded and now >= boom_until:  # Guard
            update_color_fade(now)  # Update balloon.fillColor smoothly

        # ----------------- KEYBOARD INPUT -----------------
        keys = kb.getKeys(keyList=["space", "c", "escape"], waitRelease=False, clear=True)  # Set keys

        if any(k.name == "escape" for k in keys):  # Conditional branch
            send_marker("BART_ABORT")  # Call send_marker()
            cleanup_and_exit(fh=f, send_final=False, total_bank=bank)  # Call cleanup_and_exit()

        note_text.text = "Max pumps reached. Press C to collect." if max_reached else ""  # Execute statement

        # ----------------- COLLECT (C) -----------------
        if (
            now >= cool_until
            and now >= boom_until
            and not collected
            and any(k.name == "c" for k in keys)
        ):
            collected = True

            # latency stamps for analysis
            try:
                collect_latency_from_trial_start = float(now - trial_start)
                collect_latency_from_ready = float(now - last_ready_time)
            except Exception:
                collect_latency_from_trial_start = ""
                collect_latency_from_ready = ""

            # compute earnings and update bank once
            earnings = pumps * POINTS_PER_PUMP
            bank += earnings
            events.append("collect")

            # marker + log
            send_marker(
                "BART_COLLECT",
                block=block_name,
                trial=tnum,
                pump=pumps,
                earnings=earnings,
                total=bank,
                latency_from_start=collect_latency_from_trial_start,
                latency_from_ready=collect_latency_from_ready,
            )

            # play overlay in a dedicated loop (robust on lab GPUs)
            collect_text.text = f"+{earnings} pts"
            show_dot = False
            play_collect_animation(earnings)

            # end trial with fixation
            fixation.draw()
            safe_flip()
            send_marker("BART_FIXATION_START", block=block_name, trial=tnum)
            core.wait(FIXATION_BASELINE)
            safe_flip()
            send_marker("BART_FIXATION_END", block=block_name, trial=tnum)
            break

        # ----------------- PUMP (SPACE) -----------------

        if (
            now >= cool_until  # Execute statement
            and now >= collect_until  # Execute statement
            and now >= boom_until  # Execute statement
            and not exploded  # Execute statement
            and not max_reached  # Execute statement
        ):
            pressed_space = any(k.name == "space" for k in keys)  # Set pressed_space
            if pressed_space:  # Conditional branch
                # If we already reached PUMPS_MAX, lock out further pumping
                if pumps >= PUMPS_MAX:  # Conditional branch
                    max_reached = True  # Set max_reached
                    show_dot = False  # Set show_dot
                else:  # Fallback branch
                    # Increment pump count
                    pumps += 1  # Execute statement


                    # latency from dot-ready to this pump
                    try:  # Begin protected block (handle errors)
                        lat = float(now - last_ready_time)  # Set lat
                        if lat >= 0:  # Conditional branch
                            pump_latencies.append(lat)  # Execute statement
                            pump_times.append(float(now - trial_start))  # Execute statement
                    except Exception:  # Handle an error case
                        pass  # No-op placeholder
                    # Start balloon growth animation
                    target = min(
                        balloon.radius * BALLOON_GROWTH_FACTOR + BALLOON_GROWTH_ADD,
                        BALLOON_MAX_RADIUS,
                    )
                    inflate_tween.start(balloon.radius, target, PUMP_ANIM_SEC)  # Execute statement

                    events.append(f"pump@{pumps};key=space")  # Execute statement
                    send_marker(
                        "BART_PUMP",
                        block=block_name,
                        trial=tnum,
                        pump=pumps,
                        key="space",
                    )

                    # Check whether the balloon explodes on this pump
                    if (explosion_point is not None) and (pumps >= explosion_point):  # Conditional branch
                        exploded = True  # Set exploded
                        pending_loss = pumps * POINTS_PER_PUMP  # Set pending_loss
                        bank -= pending_loss  # Execute statement
                        if not ALLOW_NEGATIVE_BANK and bank < 0:  # Conditional branch
                            bank = 0  # Set bank
                        
                        events.append(f"explode;loss={pending_loss}")
                        send_marker(
                            "BART_EXPLODE",
                            block=block_name,
                            trial=tnum,
                            pump=pumps,
                            loss=pending_loss,
                            total=bank,
                        )

                        # ERP NOTE: freeze color transitions on the explosion marker frame
                        # and render BOOM/flash in a dedicated overlay loop so it's reliably visible.
                        show_dot = False
                        play_boom_animation(pending_loss)
                        # Hard reset visuals after BOOM so next trial always starts clean
                        reset_balloon_visual()

                        # end trial with fixation
                        fixation.draw()
                        safe_flip()
                        send_marker("BART_FIXATION_START", block=block_name, trial=tnum)
                        core.wait(FIXATION_BASELINE)
                        safe_flip()
                        send_marker("BART_FIXATION_END", block=block_name, trial=tnum)
                        break

                    else:  # Fallback branch
                        # No explosion → set cooldown before next pump
                        cool_until = now + PUMP_DELAY  # Set cool_until
                        show_dot = False  # Set show_dot

        # After cooldown, re-show dot if still pumping is possible
        if (
            cool_until != 0.0  # Execute statement
            and now >= cool_until  # Execute statement
            and not exploded  # Execute statement
            and not max_reached  # Execute statement
            and (collect_until == 0.0)  # Call and()
        ):
            cool_until = 0.0  # Set cool_until
            show_dot = True  # Set show_dot

            last_ready_time = now  # Set last_ready_time
        # Hard cap guard
        if pumps >= PUMPS_MAX and not exploded:  # Conditional branch
            max_reached = True  # Set max_reached
            show_dot = False  # Set show_dot

        # ----------------- UPDATE TWEENS -----------------
        if inflate_tween.active:  # Conditional branch
            balloon.radius = inflate_tween.update()  # Execute statement
        if pop_tween.active:  # Conditional branch
            balloon.radius = pop_tween.update()  # Execute statement
        if fade_tween.active:  # Conditional branch
            balloon.opacity = fade_tween.update()  # Execute statement

        # Flash overlay for explosion
        if flash_until and now < flash_until:  # Conditional branch
            flash_rect.opacity = FLASH_OPACITY  # reduced-intensity flash  # Execute statement
        else:  # Fallback branch
            flash_rect.opacity = 0.0  # Execute statement

        # ----------------- DRAW FRAME -----------------
        # Flash overlay (drawn first so BOOM/collect text stays on top)
        if flash_rect.opacity > 0.0:
            flash_rect.draw()

        if balloon_visible:  # Conditional branch
            balloon.draw()  # Execute statement

        draw_hud(block_name, tnum, n_in_block, pumps, bank)  # Call draw_hud()

        # Dot only visible if balloon is active, no overlays, and no cooldown
        if (
            balloon_visible  # Execute statement
            and show_dot  # Execute statement
            and (boom_until == 0.0)  # Call and()
            and (collect_until == 0.0)  # Call and()
        ):
            pump_dot.draw()  # Execute statement

            current_value = pumps * POINTS_PER_PUMP  # Set current_value
            pump_value_text.text = str(current_value if current_value > 0 else "")  # Execute statement
            pump_value_text.pos = pump_dot.pos  # Execute statement
            pump_value_text.draw()  # Execute statement

        # Note text and NF HUD
        if note_text.text:  # Conditional branch
            note_text.draw()  # Execute statement
        if SHOW_NF_HUD:  # Conditional branch
            nf_status.draw()  # Execute statement
        if DEBUG_GRAPH and SHOW_NF_HUD:  # Conditional branch
            draw_debug_graph(nf)  # Call draw_debug_graph()

        # Explosion overlay
        if boom_until != 0.0 and now < boom_until and boom_t0 is not None:  # Conditional branch
            t = now - boom_t0  # Set t
            op = ease_io(t, BOOM_DUR)  # Set op
            y_main = float_up(16, 28, t, BOOM_DUR)  # Set y_main
            y_sub = float_up(-24, -10, t, BOOM_DUR)  # Set y_sub
            boom_text.opacity = op  # Execute statement
            loss_text.opacity = op  # Execute statement
            boom_text.pos = (0, y_main)  # Execute statement
            loss_text.pos = (0, y_sub)  # Execute statement
            boom_text.draw()  # Execute statement
            loss_text.draw()  # Execute statement

        # Collect overlay
        elif collect_until != 0.0 and now < collect_until and collect_t0 is not None:  # Alternative conditional branch
            t = now - collect_t0  # Set t
            op = ease_io(t, COLLECT_DUR)  # Set op
            yc = float_up(-6, 10, t, COLLECT_DUR)  # Set yc
            collect_text.opacity = op  # Execute statement
            collect_text.pos = (0, yc)  # Execute statement
            collect_text.draw()  # Execute statement

        safe_flip()  # Call safe_flip()

        # ----------------- END TRIAL CONDITIONS -----------------
        if boom_until != 0.0 and now >= boom_until:  # Conditional branch'
            reset_balloon_visual()
            fixation.draw()  # Execute statement
            safe_flip()  # Call safe_flip()
            send_marker("BART_FIXATION_START", block=block_name, trial=tnum)  # Call send_marker()
            core.wait(FIXATION_BASELINE)  # Execute statement
            safe_flip()  # Call safe_flip()
            send_marker("BART_FIXATION_END", block=block_name, trial=tnum)  # Call send_marker()
            break  # Exit current loop

        if collect_until != 0.0 and now >= collect_until and boom_until == 0.0:  # Conditional branch
            fixation.draw()  # Execute statement
            safe_flip()  # Call safe_flip()
            send_marker("BART_FIXATION_START", block=block_name, trial=tnum)  # Call send_marker()
            core.wait(FIXATION_BASELINE)  # Execute statement
            safe_flip()  # Call safe_flip()
            send_marker("BART_FIXATION_END", block=block_name, trial=tnum)  # Call send_marker()
            break  # Exit current loop

        core.wait(0.001)  # Execute statement

    # ----------------- POST-TRIAL NF METRICS -----------------
    if nf.connected and nf_frames > 0:  # Conditional branch
        nf_green_frac = nf_high_frames / float(nf_frames)  # Set nf_green_frac
        nf_green_success = nf_green_frac >= GREEN_SUCCESS_FRAC  # Set nf_green_success
        events.append(f"nf_green_frac={nf_green_frac:.2f}")  # Execute statement
    else:  # Fallback branch
        nf_green_frac = ""  # Set nf_green_frac
        nf_green_success = False  # Set nf_green_success

    t_end = core.getTime()  # Set t_end

    # ----------------- BEHAVIORAL METRICS (for offline analysis) -----------------
    exploded_int = int(bool(exploded))  # Set exploded_int
    collected_int = int(bool(collected))  # Set collected_int

    # Adjusted pumps = pumps on non-exploded trials only (standard BART metric)
    adjusted_pumps_trial = pumps if (not exploded) else ""  # Set adjusted_pumps_trial

    # Pump latency summaries (sec): dot-ready -> SPACE press
    if pump_latencies:  # Conditional branch
        pump_latency_first = float(pump_latencies[0])  # Set pump_latency_first
        pump_latency_mean = float(np.mean(pump_latencies))  # Set pump_latency_mean
        pump_latency_median = float(np.median(pump_latencies))  # Set pump_latency_median
        pump_latencies_json = json.dumps([round(x, 4) for x in pump_latencies])  # Set pump_latencies_json
        pump_times_json = json.dumps([round(x, 4) for x in pump_times])  # Set pump_times_json
    else:  # Fallback branch
        pump_latency_first = ""  # Set pump_latency_first
        pump_latency_mean = ""  # Set pump_latency_mean
        pump_latency_median = ""  # Set pump_latency_median
        pump_latencies_json = "[]"  # Set pump_latencies_json
        pump_times_json = "[]"  # Set pump_times_json
    # ----------------- WRITE CSV ROW -----------------
    row = {
            "sub": SUB_LABEL,
            "ses": SES_LABEL,
            "run": RUN_LABEL,
            "task": TASK_LABEL,
            "subject_id": SUB_LABEL,
            "name": (MANIFEST_ROW.get(MANIFEST_NAME_COL) if isinstance(MANIFEST_ROW, dict) else ""),
            "high/low": (MANIFEST_ROW.get(MANIFEST_HILO_COL) if isinstance(MANIFEST_ROW, dict) else ""),
            "condition": (CONDITION_LABEL if CONDITION_LABEL else (MANIFEST_ROW.get(MANIFEST_COND_COL) if isinstance(MANIFEST_ROW, dict) else "")),
            "block": block_name,
            "trial": tnum,
            "colour": balloon.fillColor,
            "pump_count": pumps,
            "adjusted_pumps_trial": adjusted_pumps_trial,
            "exploded_int": exploded_int,
            "collected": collected_int,
            "pump_latency_first": pump_latency_first,
            "pump_latency_mean": pump_latency_mean,
            "pump_latency_median": pump_latency_median,
            "collect_latency_from_ready": collect_latency_from_ready,
            "collect_latency_from_trial_start": collect_latency_from_trial_start,
            "pump_latencies_json": pump_latencies_json,
            "pump_times_json": pump_times_json,
            "exploded": exploded,
            "explosion_point": (explosion_point if explosion_point is not None else -1),
            "trial_value": pumps * POINTS_PER_PUMP,
            "loss_if_pop": (pumps * POINTS_PER_PUMP if exploded else 0),
            "trial_earnings": (
                pumps * POINTS_PER_PUMP  # Execute statement
                if (collect_until != 0.0 and boom_until == 0.0)  # Conditional branch
                else 0  # Execute statement
            ),
            "total_earnings": bank,
            "events": ";".join(events),
            "trial_start_time": trial_start,
            "trial_end_time": t_end,
            "trial_duration_sec": round(t_end - trial_start, 3),
            "nf_source": ("SHAM_NF" if SHAM_NF else ("SIM" if SIMULATE_NF else ("EEG" if nf.connected else "NONE"))),
            "z_used": (round(z_used, 3) if isinstance(z_used, (float, int)) else ""),
            "nf_color": (nf_cat if nf.connected else ""),
            "nf_green_frac": (round(nf_green_frac, 3) if isinstance(nf_green_frac, (float, int)) else ""),
            "nf_green_success": int(bool(nf_green_success)),
            "baseline_mu": (nf.baseline_mu if nf.baseline_mu is not None else ""),
            "baseline_sigma": (nf.baseline_sigma if nf.baseline_sigma is not None else ""),
            "baseline_n": (nf.baseline_n if getattr(nf, "baseline_done", False) else ""),
        }
    writer.writerow(row)  # Execute statement
    rows_buffer.append(dict(row))  # Execute statement
    f.flush()  # Execute statement

    send_marker("BART_ITI", block=block_name, trial=tnum)  # Call send_marker()
    core.wait(ITI)  # Execute statement
    return bank, bool(nf_green_success)  # Return value from function


# ----------------------------------------------------------------------
# BONUS OVERLAY FOR GREEN STREAK
# ----------------------------------------------------------------------

def show_bonus_overlay(current_bank: int, bonus: int):  # Define function show_bonus_overlay
    msg_main = f"+{bonus} BONUS!"  # Set msg_main
    msg_sub = f"Sustained green for {GREEN_STREAK_TARGET} balloons\nBank: {current_bank} pts"  # Set msg_sub
    t0 = core.getTime()  # Set t0
    DUR = 1.2  # Set DUR
    while True:  # Loop while condition holds
        now = core.getTime()  # Set now
        t = now - t0  # Set t
        if t >= DUR:  # Conditional branch
            break  # Exit current loop
        op = ease_io(t, DUR)  # Set op
        y_main = float_up(10, 24, t, DUR)  # Set y_main
        y_sub = float_up(-26, -10, t, DUR)  # Set y_sub

        bonus_text_main.text = msg_main  # Execute statement
        bonus_text_main.opacity = op  # Execute statement
        bonus_text_main.pos = (0, y_main)  # Execute statement

        bonus_text_sub.text = msg_sub  # Execute statement
        bonus_text_sub.opacity = op  # Execute statement
        bonus_text_sub.pos = (0, y_sub)  # Execute statement

        bonus_text_main.draw()  # Execute statement
        bonus_text_sub.draw()  # Execute statement
        safe_flip()  # Call safe_flip()
        core.wait(0.01)  # Execute statement


# ----------------------------------------------------------------------
# INSTRUCTIONS PAGES
# ----------------------------------------------------------------------

def instruction_pages():  # Define function instruction_pages
    return [
        (
            "Welcome",
            "You’ll play a balloon game where you pump up a balloon to earn points.\n"  # Execute statement
            "If the balloon pops, you lose the points from that balloon.\n\n"  # Execute statement
            "Press C to COLLECT at any time and bank your points.",
        ),
        (
            "Controls",
            "Inside the balloon you’ll see a small dot in the center.\n"  # Execute statement
            "Press the SPACE BAR to pump the balloon when the dot is visible.\n"  # Execute statement
            "There is a short cooldown after each pump, and the dot will disappear\n"  # Execute statement
            "until the balloon is ready to be pumped again.",
        ),
        (
            "Neurofeedback",
            "The balloon color reflects a measure of your brain activity in real time.\n"  # Execute statement
            "In this testing version, it may be driven by a SIMULATED signal.\n"  # Execute statement
            "  GREEN  = higher target activity\n"  # Execute statement
            "  YELLOW = medium\n"  # Execute statement
            "  RED    = lower\n\n"  # Execute statement
            "Try to keep the balloon GREEN while making good decisions.",
        ),
        (
            "Risk & Pops",
            "Each pump:\n"  # Execute statement
            "  • Adds points to this balloon’s value\n"  # Execute statement
            "  • Increases the chance the balloon will pop\n\n"  # Execute statement
            "If it pops, you lose that balloon’s value from your bank.",
        ),
        (
            "Bonus",
            f"If you keep the balloon in the GREEN state for {GREEN_STREAK_TARGET} balloons in a row,\n"  # Execute statement
            f"you earn a +{BONUS_POINTS} point bonus.\n\n"  # Execute statement
            "You’ll see a bonus message when this happens.",
        ),
        (
            "Practice & Baseline",
            "First you’ll do a few PRACTICE balloons to get used to the controls.\n"  # Execute statement
            "During these practice balloons, we also measure your brain activity (or simulated signal)\n"  # Execute statement
            "to set a baseline for the neurofeedback.\n\n"  # Execute statement
            "Then you’ll play the main game for points.\n\n"  # Execute statement
            "Press SPACE to start.",
        ),
    ]



# ---------------- COMPREHENSION CHECK ----------------
# NOTE: A brief (1–2 item) check improves data quality by ensuring participants understand the rules/keys.
def run_comprehension_check():  # short comprehension screen(s)
    send_marker('BART_COMPREHENSION_START')  # log start of check to LSL markers
    questions = [
        {
            'q': 'Question 1/2: Which key PUMPS the balloon when the dot is visible?\n\n1 = SPACE   2 = C',
            'correct': '1',
            'tag': 'pump_key',
        },
        {
            'q': "Question 2/2: What happens if the balloon POPS?\n\n1 = You lose that balloon's value from your bank\n2 = Nothing happens",
            'correct': '1',
            'tag': 'pop_rule',
        },
    ]

    for item in questions:  # loop through questions
        attempts = 0  # attempt counter per question
        while True:  # repeat until correct
            attempts += 1  # increment attempts
            title = visual.TextStim(win, text='Quick Check', pos=(0, 260), height=44, color=UI_TEXT_COLOR, bold=True)  # heading
            body  = visual.TextStim(win, text=item['q'], pos=(0, 40), height=28, color=UI_TEXT_COLOR, wrapWidth=1000, alignText='left', bold=True)  # prompt
            hint  = visual.TextStim(win, text='Press 1 or 2  (ESC to quit)', pos=(0, -300), height=22, color=UI_ACCENT_COLOR)  # instructions
            title.draw(); body.draw(); hint.draw()  # draw text
            safe_flip()  # show screen

            keys = event.waitKeys(keyList=['1', '2', 'escape'])  # wait for response
            if 'escape' in keys:  # allow abort
                send_marker('BART_ABORT')  # marker for abort
                cleanup_and_exit(fh=f, send_final=False)  # exit cleanly

            resp = keys[0]  # record response
            send_marker('BART_COMPREHENSION_RESP', item=item['tag'], resp=resp, attempts=attempts)  # log response

            if resp == item['correct']:  # correct answer
                ok = visual.TextStim(win, text='✓ Correct', height=40, color=UI_ACCENT_COLOR, bold=True)  # confirmation
                ok.draw(); safe_flip()  # show confirmation
                core.wait(0.35)  # brief pause
                break  # move to next question

            msg = visual.TextStim(win, text='Not quite — please answer again.', height=30, color=UI_TEXT_COLOR, bold=True)  # error feedback
            msg.draw(); safe_flip()  # display feedback
            core.wait(0.6)  # pause before retry

    send_marker('BART_COMPREHENSION_END')  # log end of comprehension check

def show_pages(pages):  # Define function show_pages
    send_marker(
        "BART_INSTRUCTIONS_START",
        pages=len(pages),
        practice=PRACTICE_TRIALS,
        hazard="linear",
    )
    for i, (t, b) in enumerate(pages, start=1):  # Loop over items
        title = visual.TextStim(win, text=t, pos=(0, 220), height=40, color=UI_TEXT_COLOR, bold=True)  # Set title
        body = visual.TextStim(
            win, text=b, pos=(0, 20), height=26, color=UI_TEXT_COLOR, wrapWidth=1000, alignText="left"  # Execute statement
        )
        foot = visual.TextStim(
            win, text="Press SPACE to continue", pos=(0, -300), height=22, color=UI_ACCENT_COLOR  # Execute statement
        )

        title.draw()  # Execute statement
        body.draw()  # Execute statement
        foot.draw()  # Execute statement
        safe_flip()  # Call safe_flip()

        ks = event.waitKeys(keyList=["space", "escape"])  # Set ks
        if "escape" in ks:  # Conditional branch
            cleanup_and_exit(fh=f, send_final=False)  # Call cleanup_and_exit()

        send_marker("BART_INSTRUCTIONS_PAGE", page=i, title=t)  # Call send_marker()
    send_marker("BART_INSTRUCTIONS_END")  # Call send_marker()



# ----------------------------------------------------------------------
# RESTING-STATE BLOCKS (EO/EC) FOR PRE/POST THETA + Z SUMMARY
# NOTE: Rest blocks (eyes open/closed) before & after the task to summarize resting theta and z; results are written into every trial row for easy downstream analysis.
# ----------------------------------------------------------------------

REST_SEC = 60.0  # duration per resting block (seconds)
CONC_SEC = 30.0  # duration for Concentrated Rest baseline (seconds)
REST_SAMPLE_HZ = 10.0  # how often to sample (upper bound); actual theta updates are governed by NF_UPDATE_INTERVAL

def _rest_block_screen(title: str, body: str, allow_continue=True):  # Define function _rest_block_screen
    title_stim = visual.TextStim(win, text=title, pos=(0, 220), height=40, color=UI_TEXT_COLOR, bold=True)  # Set title_stim
    # Set body_stim
    body_stim  = visual.TextStim(win, text=body,  pos=(0, 20),  height=26, color=UI_TEXT_COLOR, wrapWidth=1000, alignText="left")
    foot_stim  = visual.TextStim(win, text=("Press SPACE to begin" if allow_continue else ""),
                                 pos=(0, -300), height=22, color=UI_ACCENT_COLOR)  # Set pos
    while True:  # Loop while condition holds
        title_stim.draw()  # Execute statement
        body_stim.draw()  # Execute statement
        if allow_continue:  # Conditional branch
            foot_stim.draw()  # Execute statement
        safe_flip()  # Call safe_flip()
        ks = event.getKeys(keyList=["space","escape"])  # Set ks
        if "escape" in ks:  # Conditional branch
            send_marker("BART_ABORT")  # Call send_marker()
            cleanup_and_exit(fh=f, send_final=False)  # Call cleanup_and_exit()
        if allow_continue and "space" in ks:  # Conditional branch
            return  # Return value from function

def run_rest_block(nf: NFConnector, tag: str, duration_s: float, eyes_closed: bool):  # Define function run_rest_block
    """Run a fixation-based rest block and return summary stats for theta and z.  # Start/continue docstring

    tag: e.g., 'pre_eo', 'pre_ec', 'post_eo', 'post_ec'  # Execute statement
    """  # Start/continue docstring
    title = "Resting Block"  # Set title
    if eyes_closed:  # Conditional branch
        body = (
            "For the next minute, please keep your eyes CLOSED and relax.\n"  # Execute statement
            "Try to stay still, breathe naturally, and avoid large movements.\n\n"  # Execute statement
            "We are measuring your resting brain activity."  # Execute statement
        )
    else:  # Fallback branch
        body = (
            "For the next minute, please keep your eyes OPEN and look at the + sign.\n"  # Execute statement
            "Try to stay still, breathe naturally, and avoid large movements.\n\n"  # Execute statement
            "We are measuring your resting brain activity."  # Execute statement
        )

    _rest_block_screen(title, body, allow_continue=True)  # Call _rest_block_screen()

    send_marker("REST_START", tag=tag, eyes=("closed" if eyes_closed else "open"), dur=duration_s)  # Call send_marker()

    # countdown fixation + sampling
    t0 = core.getTime()  # Set t0
    next_sample = t0  # Set next_sample
    last_theta_time = getattr(nf, "last_theta_time", None)  # Set last_theta_time

    theta_samples = []  # Set theta_samples
    z_samples = []  # Set z_samples

    while True:  # Loop while condition holds
        now = core.getTime()  # Set now
        t = now - t0  # Set t
        if t >= duration_s:  # Conditional branch
            break  # Exit current loop

        # draw fixation and countdown
        remaining = int(duration_s - t + 0.999)  # Set remaining
        fixation.draw()  # Execute statement
        cd = visual.TextStim(win, text=f"{remaining}", pos=(0, -140), height=28, color=UI_ACCENT_COLOR)  # Set cd
        cd.draw()  # Execute statement
        safe_flip()  # Call safe_flip()

        # allow abort
        if event.getKeys(keyList=["escape"]):  # Conditional branch
            send_marker("BART_ABORT")  # Call send_marker()
            cleanup_and_exit(fh=f, send_final=False)  # Call cleanup_and_exit()

        # sample at REST_SAMPLE_HZ, but only record a sample when theta updated (real/sim/sham)
        if now >= next_sample:  # Conditional branch
            next_sample = now + (1.0 / max(1e-6, REST_SAMPLE_HZ))  # Set next_sample
            z = nf.pull_z()  # Set z
            # record if theta updated since last record
            th_t = getattr(nf, "last_theta_time", None)  # Set th_t
            th = getattr(nf, "last_theta", None)  # Set th
            if th is not None and th_t is not None and th_t != last_theta_time:  # Conditional branch
                last_theta_time = th_t  # Set last_theta_time
                theta_samples.append(float(th))  # Execute statement
                z_samples.append(float(z))  # Execute statement

        core.wait(0.001)  # Execute statement

    send_marker("REST_END", tag=tag, eyes=("closed" if eyes_closed else "open"))  # Call send_marker()


    # Play an end-of-block chime ONLY after eyes-closed rest (EC).
    if eyes_closed and CHIME_ENABLED and CHIME_OK and (end_chime is not None):  # Conditional branch
        try:  # Begin protected block (handle errors)
            send_marker("REST_EC_CHIME", tag=tag)  # Call send_marker()
            end_chime.play()  # Execute statement
            core.wait(END_CHIME_DUR + 0.05)  # Execute statement
        except Exception as _e:  # Handle an error case
            print("⚠️ Chime playback failed:", _e)  # Print debug/status message

    # summaries
    def _summ(v):  # Define function _summ
        if not v:  # Conditional branch
            return ("", "")  # Return value from function
        arr = np.asarray(v, dtype=float)  # Set arr
        m = float(np.mean(arr))  # Set m
        s = float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0  # Set s
        return m, s  # Return value from function

    theta_mean, theta_std = _summ(theta_samples)  # Execute statement
    z_mean, z_std = _summ(z_samples)  # Execute statement
    n = len(theta_samples)  # Set n

    return {
        "theta_mean": theta_mean,
        "theta_std": theta_std,
        "z_mean": z_mean,
        "z_std": z_std,
        "n": n,
        "theta_samples": theta_samples,
        "z_samples": z_samples,
    }



# ----------------------------------------------------------------------
# CONCENTRATED REST BASELINE BLOCK
# ----------------------------------------------------------------------

def run_concentrated_rest(nf, block_code: str, duration_s: float):  # Define function run_concentrated_rest
    """Concentration baseline: count backwards by 7s for a short period.

    We collect theta power samples during this window so we can *calibrate the
    direction* of neurofeedback (some participants show inverted patterns).

    Returns the same metric dict as run_rest_block, including theta_samples.
    """  # End docstring

    # Create a random 3-digit start number for the mental arithmetic prompt
    start_num = random.randint(100, 999)  # Random start value

    # User-friendly instruction text
    title = visual.TextStim(win, text="Concentrated Rest", height=44, color=UI_TEXT_COLOR, pos=(0, 200), bold=True)
    body  = visual.TextStim(
        win,
        text=("""A 3-digit number will appear.

Please count backwards out loud or silently by 7s (e.g., 392, 385, 378, ...).
This is a short concentration rest.

This helps us calibrate the neurofeedback baseline."""),
        height=26,
        color=UI_TEXT_COLOR,
        wrapWidth=1000,
        pos=(0, 10)
    )
    prompt = visual.TextStim(win, text="Press SPACE to begin", height=24, color=UI_ACCENT_COLOR, pos=(0, -260))

    # Show page and wait
    title.draw(); body.draw(); prompt.draw(); safe_flip()
    ks = event.waitKeys(keyList=["space", "escape"])  # Wait for keypress
    if "escape" in ks:
        send_marker("BART_ABORT")
        cleanup_and_exit(fh=f, send_final=False)

    # Countdown for timing clarity
    send_marker("REST_START", block=block_code, kind="concentrated", dur=duration_s)
    countdown = visual.TextStim(win, text="3", height=72, color=UI_TEXT_COLOR, pos=(0, 0), bold=True)
    for n in ["3", "2", "1"]:
        countdown.text = n
        countdown.draw(); safe_flip(); core.wait(1.0)

    # Display the number + fixation during the block
    num_stim = visual.TextStim(win, text=str(start_num), height=72, color=UI_TEXT_COLOR, pos=(0, 40), bold=True)
    instr    = visual.TextStim(win, text="Count backwards by 7s", height=28, color=UI_TEXT_COLOR, pos=(0, -60))

    # Sample theta/z during this block (same sampling method as run_rest_block)
    t0 = core.getTime()
    next_sample = t0
    theta_vals = []
    z_vals = []
    theta_samples = []
    z_samples = []

    while core.getTime() - t0 < duration_s:
        # update NF; pull_z updates nf.last_theta and returns z
        z = nf.pull_z()
        theta = nf.last_theta

        # sample at REST_SAMPLE_HZ
        now = core.getTime()
        if now >= next_sample:
            if theta is not None:
                theta_vals.append(theta)
                theta_samples.append(theta)
            if z is not None:
                z_vals.append(z)
                z_samples.append(z)
            next_sample = now + (1.0 / max(1e-6, REST_SAMPLE_HZ))

        # draw frame
        num_stim.draw()
        instr.draw()
        safe_flip()
        core.wait(0.001)

    send_marker("REST_END", block=block_code, kind="concentrated")

    # Compute summary stats
    def _mean_std(arr):
        arr = np.asarray(arr, float)
        arr = arr[np.isfinite(arr)]
        if len(arr) == 0:
            return ("", "", 0)
        return (float(arr.mean()), float(arr.std(ddof=1)) if len(arr) > 1 else 0.0, int(len(arr)))

    theta_mean, theta_std, n = _mean_std(theta_vals)
    z_mean, z_std, _ = _mean_std(z_vals)

    return {
        "theta_mean": theta_mean,
        "theta_std": theta_std,
        "z_mean": z_mean,
        "z_std": z_std,
        "n": n,
        "theta_samples": theta_samples,
        "z_samples": z_samples,
    }
def apply_rest_columns_to_rows(rows, rest_metrics: dict):  # Define function apply_rest_columns_to_rows
    """Fill rest_* columns on every trial row."""  # Start/continue docstring
    for r in rows:  # Loop over items
        for rb in ["pre_eo","pre_ec","pre_conc","post_eo","post_ec"]:  # Loop over items
            met = rest_metrics.get(rb, {})  # Set met
            r[f"rest_{rb}_theta_mean"] = met.get("theta_mean", "")  # Execute statement
            r[f"rest_{rb}_theta_std"]  = met.get("theta_std", "")  # Execute statement
            r[f"rest_{rb}_z_mean"]     = met.get("z_mean", "")  # Execute statement
            r[f"rest_{rb}_z_std"]      = met.get("z_std", "")  # Execute statement
            r[f"rest_{rb}_n"]          = met.get("n", "")  # Execute statement
    return rows  # Return value from function

def rewrite_csv_from_buffer(csv_path: str, fieldnames: list, rows: list):  # Define function rewrite_csv_from_buffer
    """Rewrite CSV entirely from rows buffer (ensures new columns are populated)."""  # Start/continue docstring
    try:  # Begin protected block (handle errors)
        with open(csv_path, "w", newline="") as _fh:  # Context manager (auto cleanup)
            w = csv.DictWriter(_fh, fieldnames=fieldnames)  # Set w
            w.writeheader()  # Execute statement
            for r in rows:  # Loop over items
                w.writerow({k: r.get(k, "") for k in fieldnames})  # Execute statement
    except Exception as e:  # Handle an error case
        print("⚠️ Could not rewrite CSV with rest columns:", e)  # Print debug/status message


# ----------------------------------------------------------------------
# MAIN SCRIPT ENTRY POINT
# ----------------------------------------------------------------------

nf = NFConnector()  # Set nf

# ----------------- CONNECT NF / EEG -----------------
if REQUIRE_NF and not SIMULATE_NF:  # Conditional branch
    wait_txt = visual.TextStim(
        win,
        text=f"Waiting for EEG stream '{EEG_STREAM_NAME}'...\n(Press ESC to quit)",
        height=28,
        color=UI_TEXT_COLOR,
    )
    while not nf.connected:  # Loop while condition holds
        wait_txt.draw()  # Execute statement
        safe_flip()  # Call safe_flip()
        if event.getKeys(keyList=["escape"]):  # Conditional branch
            cleanup_and_exit(fh=f, send_final=False)  # Call cleanup_and_exit()
        nf.try_connect(attempts=1, sleep_s=0.5)  # Execute statement
        core.wait(0.5)  # Execute statement
else:  # Fallback branch
    nf.try_connect(attempts=10, sleep_s=0.2)  # Execute statement


# ----------------- REST OVERVIEW SCREEN -----------------
# --- REST INSTRUCTIONS SCREEN ---
title = visual.TextStim(
    win,
    text="Resting Baseline",
    height=40,
    pos=(0, 220),
    color=UI_TEXT_COLOR,
    bold=True  # Set bold
)

body = visual.TextStim(
    win,
    text=(
        "Before the task begins, you will complete two short resting periods.\n\n"  # Execute statement
        "• Eyes OPEN (1 minute):\n"  # Execute statement
        "  Look at the fixation cross and relax.\n\n"  # Execute statement
        "• Eyes CLOSED (1 minute):\n"  # Execute statement
        "  Close your eyes and stay relaxed.\n\n"  # Execute statement
        "These periods allow us to measure your resting brain activity.\n\n"  # Execute statement
        "Press SPACE to begin."  # Execute statement
    ),
    height=26,
    wrapWidth=1000,
    pos=(0, 0),
    color=UI_TEXT_COLOR,
    alignText="left"  # Set alignText
)

footer = visual.TextStim(
    win,
    text="Press SPACE to continue",
    height=22,
    pos=(0, -300),
    color=UI_ACCENT_COLOR  # Set color
)

while True:  # Loop while condition holds
    title.draw()  # Execute statement
    body.draw()  # Execute statement
    footer.draw()  # Execute statement
    win.flip()  # Execute statement

    keys = event.getKeys(keyList=["space", "escape"])  # Set keys
    if "escape" in keys:  # Conditional branch
        cleanup_and_exit(fh=f, send_final=False)  # Call cleanup_and_exit()
    if "space" in keys:  # Conditional branch
        break  # Exit current loop


# ----------------- PRE REST (EO/EC) -----------------
rest_metrics = {}  # Set rest_metrics
send_marker("BART_BASELINE_START", phase="rest_calibration")  # baseline calibration begins
rest_metrics["pre_eo"] = run_rest_block(nf, "pre_eo", REST_SEC, eyes_closed=False)  # Execute statement
rest_metrics["pre_ec"] = run_rest_block(nf, "pre_ec", REST_SEC, eyes_closed=True)  # Execute statement

# --- Concentrated Rest (baseline direction calibration) ---
rest_metrics["pre_conc"] = run_concentrated_rest(nf, "pre_conc", CONC_SEC)  # collect concentration samples

# --- Compute NF baseline from EO rest + concentration block ---
nf.set_baseline_from_rest_epochs(
    rest_metrics.get("pre_eo", {}).get("theta_samples", []),
    rest_metrics.get("pre_conc", {}).get("theta_samples", [])
)
send_marker("BART_BASELINE_END", phase="rest_calibration", mu=nf.baseline_mu, sigma=nf.baseline_sigma, dir=nf.baseline_direction, n=nf.baseline_n)

# ----------------- INSTRUCTIONS -----------------
pages = instruction_pages()  # Set pages
show_pages(pages)  # Call show_pages()

# ----------------- COMPREHENSION CHECK -----------------  # quick check to ensure participant understands task rules
run_comprehension_check()  # ask 2 short questions (pump key + pop consequence) before starting the task


# Show NF source
splash = visual.TextStim(
    win,
    text=f"NF source: {'EEG/Sim/Sham' if (SIMULATE_NF or SHAM_NF or nf.connected) else 'NONE'}\n(Press any key to continue)",
    height=28,
    color=UI_TEXT_COLOR,
)
splash.draw()  # Execute statement
safe_flip()  # Call safe_flip()
event.waitKeys()  # Execute statement



# ----------------- PRACTICE (LEARNING ONLY) -----------------
# Practice trials help participants learn the game mechanics.
# We do NOT compute the NF baseline here anymore (baseline is from rest blocks).

total_bank = 0  # reset bank for practice
send_marker("BART_PRACTICE_START")  # mark practice start

for tr in practice_trials:  # loop practice balloons
    total_bank, _ = run_trial(
        "Practice",
        0,
        tr,
        PRACTICE_TRIALS,
        total_bank,
        nf=nf,
        use_fallback_colors=False,
        nf_color_enabled=False,
    )

send_marker("BART_PRACTICE_END")  # mark practice end

# Reset bank for main
total_bank = 0  # Set total_bank
btw = visual.TextStim(
    win,
    text="Practice complete.\nYour bank is now reset to 0.\n\nPress any key for the MAIN game.",
    height=28,
    color=UI_TEXT_COLOR,
)
btw.draw()  # Execute statement
safe_flip()  # Call safe_flip()
event.waitKeys()  # Execute statement

# ----------------- MAIN TASK -----------------
green_streak = 0  # Set green_streak

for tr in main_trials:  # Loop over items
    total_bank, green_success = run_trial(
        "Main",
        1,
        tr,
        N_TRIALS,
        total_bank,
        nf=nf,
        use_fallback_colors=False,
    )

    if green_success:  # Conditional branch
        green_streak += 1  # Execute statement
    else:  # Fallback branch
        green_streak = 0  # Set green_streak

    if green_streak >= GREEN_STREAK_TARGET:  # Conditional branch
        total_bank += BONUS_POINTS  # Execute statement
        send_marker(
            "NF_BONUS",
            block="Main",
            bonus=BONUS_POINTS,
            total=total_bank,
            streak=green_streak,
        )
        show_bonus_overlay(total_bank, BONUS_POINTS)  # Call show_bonus_overlay()
        green_streak = 0  # Set green_streak


# ----------------- MAIN COMPLETE → POST-REST TRANSITION -----------------
# NOTE: Participants must intentionally press SPACE to begin the post-task resting blocks.
# This prevents an abrupt jump from the 'finished' feeling into EO/EC rest without warning.
post_rest_prompt = visual.TextStim(
    win,
    text=(
        "Nice work — the game is complete!\n\n"
        "Next you will do two short resting blocks to measure resting theta:\n"
        "  • 1 minute Eyes Open (look at the +)\n"
        "  • 1 minute Eyes Closed (you’ll hear a chime when it’s done)\n\n"
        "Press SPACE when you’re ready to continue."
    ),
    height=28,
    color=UI_TEXT_COLOR,
    wrapWidth=1000,
)
post_rest_prompt.draw()
safe_flip()
ks = event.waitKeys(keyList=['space','escape'])
if ks and ('escape' in ks):
    send_marker('BART_ABORT', reason='escape_at_post_rest_prompt')
    cleanup_and_exit(fh=f, send_final=False, total_bank=total_bank)
# ----------------- POST REST (EO/EC) -----------------
rest_metrics["post_eo"] = run_rest_block(nf, "post_eo", REST_SEC, eyes_closed=False)  # Execute statement
rest_metrics["post_ec"] = run_rest_block(nf, "post_ec", REST_SEC, eyes_closed=True)  # Execute statement

# Apply rest columns to all rows and rewrite CSV so rest values are present
apply_rest_columns_to_rows(rows_buffer, rest_metrics)  # Call apply_rest_columns_to_rows()
rewrite_csv_from_buffer(csvfile, FIELDNAMES, rows_buffer)  # Call rewrite_csv_from_buffer()

# ----------------- END SCREEN -----------------
end_text = visual.TextStim(
    win,
    text=f"All done!\n\nFinal bank: {total_bank} points\n\nThank you for playing.\n\nPress any key to exit.",
    height=30,
    color=UI_TEXT_COLOR,
)
end_text.draw()  # Execute statement
safe_flip()  # Call safe_flip()
event.waitKeys()  # Execute statement

cleanup_and_exit(fh=f, send_final=True, total_bank=total_bank)  # Call cleanup_and_exit()